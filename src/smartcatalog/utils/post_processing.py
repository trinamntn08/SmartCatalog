from __future__ import annotations

from dataclasses import dataclass, field
import io
import math
from pathlib import Path, PurePosixPath
import posixpath
import re
from typing import Optional
import xml.etree.ElementTree as ET
import zipfile

from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PILImage


PX_TO_EMU = 9525
DEFAULT_SHEET_NAME = "form đầu ra"
DEFAULT_BG_RELATIVE_PATH = Path("icons") / "bg.jpg"
_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_DOC_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"


@dataclass(slots=True)
class PostProcessingRow:
    number: int
    code: str
    qty: str = ""
    desc_primary: str = ""
    desc_secondary: str = ""
    image_paths: list[str] = field(default_factory=list)
    highlight_missing_desc: bool = False
    force_secondary_row: bool = False
    highlight_missing_secondary: bool = False


def resolve_post_processing_bg_path(project_dir: str | Path) -> Optional[Path]:
    root = Path(project_dir).resolve()
    candidate = root / DEFAULT_BG_RELATIVE_PATH
    return candidate if candidate.is_file() else None


def apply_post_processing_branding(
    xlsx_path: str | Path,
    *,
    project_dir: str | Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> bool:
    bg_path = resolve_post_processing_bg_path(project_dir)
    if not bg_path:
        return False
    _inject_header_image(Path(xlsx_path), sheet_name=sheet_name, bg_path=bg_path)
    return True


def write_post_processing_sheet(
    workbook,
    rows: list[PostProcessingRow],
    *,
    sheet_name: str = DEFAULT_SHEET_NAME,
    preserve_image_order: bool = False,
    index: Optional[int] = None,
) -> Worksheet:
    sheet_index = index
    if sheet_name in workbook.sheetnames:
        existing = workbook[sheet_name]
        existing_index = workbook.worksheets.index(existing)
        workbook.remove(existing)
        if sheet_index is not None and existing_index < sheet_index:
            sheet_index -= 1

    if sheet_index is None:
        ws = workbook.create_sheet(sheet_name)
    else:
        ws = workbook.create_sheet(sheet_name, max(0, sheet_index))

    _configure_sheet(ws)
    _write_header(ws)

    row_idx = 2
    for item in rows:
        row_idx = _write_product_block(
            ws,
            row_idx,
            item,
            preserve_image_order=preserve_image_order,
        )

    return ws


def _configure_sheet(ws: Worksheet) -> None:
    ws.column_dimensions["A"].width = 9.11
    ws.column_dimensions["B"].width = 19.44
    ws.column_dimensions["C"].width = 121.11
    ws.column_dimensions["D"].width = 15.44
    ws.freeze_panes = "A2"
    ws.oddFooter.right.text = '&"Times New Roman,Italic"&12 Page &P of &N'
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddHeader.center.text = "&G"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "1:1"
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.8
    ws.page_margins.bottom = 0.8
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3


def _write_header(ws: Worksheet) -> None:
    header_font = Font(name="Times New Roman", size=14, bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="2F5597")
    thin = Side(style="thin", color="000000")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ["No.", "Product Code", "Product Description", "Qty."]
    for col_idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center if col_idx in (1, 2, 4) else left
        cell.border = Border(bottom=thin)
    ws.row_dimensions[1].height = 28


def _write_product_block(
    ws: Worksheet,
    start_row: int,
    item: PostProcessingRow,
    *,
    preserve_image_order: bool,
) -> int:
    data_font = Font(name="Times New Roman", size=12, bold=True)
    desc_font = Font(name="Times New Roman", size=12, bold=False)
    thin = Side(style="thin", color="000000")
    center = Alignment(horizontal="center", vertical="center")
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    yellow = PatternFill("solid", fgColor="FFFF00")

    primary_row = start_row
    has_secondary_row = bool(item.desc_secondary or item.force_secondary_row)
    secondary_row = start_row + 1 if has_secondary_row else None
    image_row = start_row + 1 + (1 if has_secondary_row else 0)

    for col_idx in range(1, 5):
        ws.cell(primary_row, col_idx).border = Border(top=thin)

    ws.cell(primary_row, 1, item.number).font = data_font
    ws.cell(primary_row, 1).alignment = center
    ws.cell(primary_row, 2, item.code).font = data_font
    ws.cell(primary_row, 2).alignment = center_wrap
    ws.cell(primary_row, 3, item.desc_primary or "").font = data_font
    ws.cell(primary_row, 3).alignment = left
    ws.cell(primary_row, 4, item.qty).font = data_font
    ws.cell(primary_row, 4).alignment = center_wrap

    if item.highlight_missing_desc and not item.desc_primary:
        ws.cell(primary_row, 3).fill = yellow

    ws.row_dimensions[primary_row].height = _row_height_for_text(item.desc_primary)

    if secondary_row is not None:
        ws.cell(secondary_row, 3, item.desc_secondary or "").font = desc_font
        ws.cell(secondary_row, 3).alignment = left
        if item.highlight_missing_secondary and not item.desc_secondary:
            ws.cell(secondary_row, 3).fill = yellow
        ws.row_dimensions[secondary_row].height = _row_height_for_text(item.desc_secondary)

    ws.merge_cells(start_row=image_row, start_column=1, end_row=image_row, end_column=4)
    ws.row_dimensions[image_row].height = 15

    _place_images(
        ws,
        image_row,
        item.image_paths,
        preserve_image_order=preserve_image_order,
    )

    return image_row + 1


def _row_height_for_text(text: str) -> float:
    raw = str(text or "").strip()
    if not raw:
        return 22.5
    lines = max(1, raw.count("\n") + 1)
    wrapped_lines = max(1, math.ceil(len(raw) / 96))
    return max(22.5, min(90.0, max(lines, wrapped_lines) * 18.0))


def _place_images(
    ws: Worksheet,
    row_idx: int,
    image_paths: list[str],
    *,
    preserve_image_order: bool,
) -> None:
    loaded = _load_images(image_paths, preserve_image_order=preserve_image_order)
    if not loaded:
        return

    available_width = sum(_col_width_px(ws, col_idx) for col_idx in range(1, 5))
    gap_px = 12
    max_row_height_px = 180

    raw_total_width = sum(w for _, w, _ in loaded) + gap_px * max(0, len(loaded) - 1)
    raw_max_height = max(h for _, _, h in loaded)
    if raw_total_width <= 0 or raw_max_height <= 0:
        return

    scale_h = min(1.0, max_row_height_px / float(raw_max_height))
    scale_w = min(1.0, available_width / float(raw_total_width)) if available_width > 0 else 1.0
    scale = min(1.0, scale_h, scale_w)

    scaled_items: list[tuple[PILImage.Image, int, int]] = []
    for pil, width, height in loaded:
        new_width = max(1, int(width * scale))
        new_height = max(1, int(height * scale))
        scaled_items.append((pil, new_width, new_height))

    band_height = max(h for _, _, h in scaled_items)
    total_width = sum(w for _, w, _ in scaled_items) + gap_px * max(0, len(scaled_items) - 1)
    start_x = max(0, int((available_width - total_width) / 2))
    v_pad = 10
    ws.row_dimensions[row_idx].height = max(15.0, (band_height + (v_pad * 2)) * 0.75)

    col_widths = [_col_width_px(ws, col_idx) for col_idx in range(1, 5)]
    x_off = start_x
    for pil, width, height in scaled_items:
        try:
            buf = io.BytesIO()
            if pil.width != width or pil.height != height:
                render = pil.resize((width, height), PILImage.LANCZOS)
            else:
                render = pil
            render.save(buf, format="PNG")
            buf.seek(0)

            xl_img = XLImage(buf)
            xl_img.width = width
            xl_img.height = height

            col_idx = 0
            col_off = x_off
            while col_idx < len(col_widths) - 1 and col_off >= col_widths[col_idx]:
                col_off -= col_widths[col_idx]
                col_idx += 1

            row_height_px = int((ws.row_dimensions[row_idx].height or 0) / 0.75)
            y_off = v_pad
            if row_height_px > height + (v_pad * 2):
                y_off = int((row_height_px - height) / 2)

            marker = AnchorMarker(
                col=col_idx,
                colOff=int(col_off * PX_TO_EMU),
                row=row_idx - 1,
                rowOff=int(y_off * PX_TO_EMU),
            )
            ext = XDRPositiveSize2D(width * PX_TO_EMU, height * PX_TO_EMU)
            xl_img.anchor = OneCellAnchor(_from=marker, ext=ext)
            ws.add_image(xl_img)

            x_off += width + gap_px
        except Exception:
            continue


def _load_images(
    image_paths: list[str],
    *,
    preserve_image_order: bool,
) -> list[tuple[PILImage.Image, int, int]]:
    loaded: list[tuple[PILImage.Image, int, int]] = []
    for raw_path in image_paths:
        try:
            path = Path(raw_path)
            if not path.exists():
                continue
            pil = PILImage.open(path).convert("RGBA")
            if not preserve_image_order and pil.height > pil.width:
                pil = pil.rotate(90, expand=True)
            width, height = pil.size
            if width <= 0 or height <= 0:
                continue
            loaded.append((pil, width, height))
        except Exception:
            continue

    if not preserve_image_order:
        loaded.sort(key=lambda item: item[1] * item[2])
    return loaded


def _col_width_px(ws: Worksheet, col_idx: int) -> int:
    letter = get_column_letter(col_idx)
    width = ws.column_dimensions[letter].width
    if width is None:
        width = 8.43
    return int(width * 7 + 5)


def _inject_header_image(xlsx_path: Path, *, sheet_name: str, bg_path: Path) -> None:
    bg_bytes = _bg_to_jpeg_bytes(bg_path)
    temp_path = xlsx_path.with_name(f"{xlsx_path.stem}.tmp{xlsx_path.suffix}")

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        names = zin.namelist()
        sheet_path = _resolve_sheet_xml_path(zin, sheet_name)
        sheet_rels_path = f"xl/worksheets/_rels/{PurePosixPath(sheet_path).name}.rels"

        old_vml_paths, old_bg_media_paths = _collect_old_header_assets(zin, sheet_path, sheet_rels_path)

        content_types = zin.read("[Content_Types].xml").decode("utf-8")
        sheet_xml = zin.read(sheet_path).decode("utf-8")
        sheet_rels_xml = (
            zin.read(sheet_rels_path).decode("utf-8")
            if sheet_rels_path in names
            else _empty_relationships_xml()
        )

        next_image_num = _next_suffix_number(names, r"xl/media/image(\d+)\.")
        bg_media_path = f"xl/media/image{next_image_num}.jpeg"

        next_vml_num = _next_suffix_number(names, r"xl/drawings/vmlDrawing(\d+)\.vml$")
        vml_name = f"vmlDrawing{next_vml_num}.vml"
        vml_path = f"xl/drawings/{vml_name}"
        vml_rels_path = f"xl/drawings/_rels/{vml_name}.rels"

        sheet_rels_xml, vml_rel_id = _upsert_vml_relationship(sheet_rels_xml, vml_name)
        sheet_xml = _upsert_header_footer_xml(sheet_xml, vml_rel_id)
        content_types = _ensure_content_type_defaults(content_types)
        vml_xml = _build_vml_xml()
        vml_rels_xml = _build_vml_rels_xml(bg_media_path)

        skip_paths = set(old_vml_paths) | set(old_bg_media_paths)
        with zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            written: set[str] = set()
            for info in zin.infolist():
                if info.filename in skip_paths:
                    continue
                data = zin.read(info.filename)
                if info.filename == sheet_path:
                    data = sheet_xml.encode("utf-8")
                elif info.filename == sheet_rels_path:
                    data = sheet_rels_xml.encode("utf-8")
                elif info.filename == "[Content_Types].xml":
                    data = content_types.encode("utf-8")
                zout.writestr(info, data)
                written.add(info.filename)

            if sheet_rels_path not in written:
                zout.writestr(sheet_rels_path, sheet_rels_xml.encode("utf-8"))
            zout.writestr(vml_path, vml_xml.encode("utf-8"))
            zout.writestr(vml_rels_path, vml_rels_xml.encode("utf-8"))
            zout.writestr(bg_media_path, bg_bytes)

    temp_path.replace(xlsx_path)


def _bg_to_jpeg_bytes(bg_path: Path) -> bytes:
    with PILImage.open(bg_path).convert("RGB") as bg:
        buf = io.BytesIO()
        bg.save(buf, format="JPEG", quality=90)
        return buf.getvalue()


def _resolve_sheet_xml_path(zin: zipfile.ZipFile, sheet_name: str) -> str:
    workbook_root = ET.fromstring(zin.read("xl/workbook.xml"))
    rels_root = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
    rel_map = {
        rel.get("Id"): rel.get("Target", "")
        for rel in rels_root.findall(f"{{{_NS_PKG_REL}}}Relationship")
    }

    for sheet in workbook_root.findall(f".//{{{_NS_MAIN}}}sheet"):
        if sheet.get("name") != sheet_name:
            continue
        rid = sheet.get(f"{{{_NS_DOC_REL}}}id")
        target = rel_map.get(rid or "", "")
        if not target:
            break
        if target.startswith("/"):
            normalized = posixpath.normpath(target.lstrip("/"))
        else:
            normalized = posixpath.normpath(posixpath.join("xl", target))
        return normalized
    raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")


def _collect_old_header_assets(
    zin: zipfile.ZipFile,
    sheet_path: str,
    sheet_rels_path: str,
) -> tuple[list[str], list[str]]:
    if sheet_rels_path not in zin.namelist():
        return [], []

    root = ET.fromstring(zin.read(sheet_rels_path))
    old_vml_paths: list[str] = []
    old_bg_media_paths: list[str] = []
    sheet_base = posixpath.dirname(sheet_path)
    for rel in root.findall(f"{{{_NS_PKG_REL}}}Relationship"):
        if rel.get("Type") != f"{_NS_DOC_REL}/vmlDrawing":
            continue
        target = rel.get("Target", "")
        if not target:
            continue
        vml_path = posixpath.normpath(posixpath.join(sheet_base, target))
        old_vml_paths.append(vml_path)

        vml_rels_path = f"xl/drawings/_rels/{PurePosixPath(vml_path).name}.rels"
        if vml_rels_path not in zin.namelist():
            continue
        rels_root = ET.fromstring(zin.read(vml_rels_path))
        vml_base = posixpath.dirname(vml_path)
        for vml_rel in rels_root.findall(f"{{{_NS_PKG_REL}}}Relationship"):
            media_target = vml_rel.get("Target", "")
            if not media_target:
                continue
            old_bg_media_paths.append(posixpath.normpath(posixpath.join(vml_base, media_target)))
        old_vml_paths.append(vml_rels_path)
    return old_vml_paths, old_bg_media_paths


def _empty_relationships_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"></Relationships>'
    )


def _next_suffix_number(names: list[str], pattern: str) -> int:
    values = [int(m.group(1)) for name in names if (m := re.match(pattern, name))]
    return max(values, default=0) + 1


def _upsert_vml_relationship(sheet_rels_xml: str, vml_name: str) -> tuple[str, str]:
    root = ET.fromstring(sheet_rels_xml)
    for rel in list(root.findall(f"{{{_NS_PKG_REL}}}Relationship")):
        if rel.get("Type") == f"{_NS_DOC_REL}/vmlDrawing":
            root.remove(rel)

    existing_ids = []
    for rel in root.findall(f"{{{_NS_PKG_REL}}}Relationship"):
        raw_id = rel.get("Id", "")
        if raw_id.startswith("rId"):
            try:
                existing_ids.append(int(raw_id[3:]))
            except ValueError:
                pass
    next_id = f"rId{max(existing_ids, default=0) + 1}"
    ET.SubElement(
        root,
        f"{{{_NS_PKG_REL}}}Relationship",
        {
            "Id": next_id,
            "Type": f"{_NS_DOC_REL}/vmlDrawing",
            "Target": f"../drawings/{vml_name}",
        },
    )
    return ET.tostring(root, encoding="unicode"), next_id


def _upsert_header_footer_xml(sheet_xml: str, vml_rel_id: str) -> str:
    legacy_pattern = re.compile(r"<legacyDrawingHF\b[^>]*/>", flags=re.S)
    sheet_xml = legacy_pattern.sub("", sheet_xml)

    if "<headerFooter" in sheet_xml:
        sheet_xml = re.sub(
            r"<oddHeader>.*?</oddHeader>",
            "<oddHeader>&amp;C&amp;G</oddHeader>",
            sheet_xml,
            count=1,
            flags=re.S,
        )
        if "<oddHeader>" not in sheet_xml:
            sheet_xml = re.sub(
                r"(<headerFooter\b[^>]*>)",
                r"\1<oddHeader>&amp;C&amp;G</oddHeader>",
                sheet_xml,
                count=1,
                flags=re.S,
            )
    else:
        sheet_xml = sheet_xml.replace(
            "</worksheet>",
            "<headerFooter><oddHeader>&amp;C&amp;G</oddHeader></headerFooter></worksheet>",
            1,
        )

    return sheet_xml.replace(
        "</worksheet>",
        f'<legacyDrawingHF xmlns:r="{_NS_DOC_REL}" r:id="{vml_rel_id}"/></worksheet>',
        1,
    )


def _ensure_content_type_defaults(content_types_xml: str) -> str:
    if 'Extension="vml"' not in content_types_xml:
        content_types_xml = content_types_xml.replace(
            "</Types>",
            '<Default Extension="vml" ContentType="application/vnd.openxmlformats-officedocument.vmlDrawing"/></Types>',
            1,
        )
    if 'Extension="jpeg"' not in content_types_xml:
        content_types_xml = content_types_xml.replace(
            "</Types>",
            '<Default Extension="jpeg" ContentType="image/jpeg"/></Types>',
            1,
        )
    return content_types_xml


def _build_vml_xml() -> str:
    return (
        '<xml xmlns:v="urn:schemas-microsoft-com:vml" '
        'xmlns:o="urn:schemas-microsoft-com:office:office" '
        'xmlns:x="urn:schemas-microsoft-com:office:excel">'
        '<o:shapelayout v:ext="edit"><o:idmap v:ext="edit" data="1"/></o:shapelayout>'
        '<v:shapetype id="_x0000_t75" coordsize="21600,21600" o:spt="75" '
        'o:preferrelative="t" path="m@4@5l@4@11@9@11@9@5xe" filled="f" stroked="f">'
        '<v:stroke joinstyle="miter"/><v:formulas>'
        '<v:f eqn="if lineDrawn pixelLineWidth 0"/><v:f eqn="sum @0 1 0"/>'
        '<v:f eqn="sum 0 0 @1"/><v:f eqn="prod @2 1 2"/>'
        '<v:f eqn="prod @3 21600 pixelWidth"/><v:f eqn="prod @3 21600 pixelHeight"/>'
        '<v:f eqn="sum @0 0 1"/><v:f eqn="prod @6 1 2"/>'
        '<v:f eqn="prod @7 21600 pixelWidth"/><v:f eqn="sum @8 21600 0"/>'
        '<v:f eqn="prod @7 21600 pixelHeight"/><v:f eqn="sum @10 21600 0"/>'
        '</v:formulas><v:path o:extrusionok="f" gradientshapeok="t" o:connecttype="rect"/>'
        '<o:lock v:ext="edit" aspectratio="t"/></v:shapetype>'
        '<v:shape id="CH" o:spid="_x0000_s1027" type="#_x0000_t75" '
        "style='position:absolute;margin-left:0;margin-top:0;width:873pt;height:1382.4pt;z-index:1' "
        'o:preferrelative="f"><v:imagedata o:relid="rId1" o:title="post_processing_bg"/>'
        '<o:lock v:ext="edit" rotation="t" aspectratio="f"/></v:shape></xml>'
    )


def _build_vml_rels_xml(bg_media_path: str) -> str:
    image_name = PurePosixPath(bg_media_path).name
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f'<Relationship Id="rId1" Type="{_NS_DOC_REL}/image" Target="../media/{image_name}"/>'
        "</Relationships>"
    )
