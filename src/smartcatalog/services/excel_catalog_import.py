from __future__ import annotations

import io
from bisect import bisect_right
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from smartcatalog.loader.excel_loader import (
    detect_excel_code_column,
    load_code_to_vi_en_from_excel,
)
from smartcatalog.utils.code_normalization import (
    build_unique_normalized_code_index,
    normalize_code_soft,
)
from smartcatalog.utils.filename_utils import sanitize_filename
from smartcatalog.utils.hashing import sha256_bytes
from smartcatalog.utils.text_normalization import normalize_header_text
from smartcatalog.utils.workbook_images import get_image_anchor_row, image_to_pil


@dataclass(frozen=True)
class ExcelCatalogImportResult:
    total: int
    created_new: int
    updated: int
    skipped_existing: int
    missing: int
    missing_codes: tuple[str, ...] = ()
    images_updated: int = 0
    images_missing: int = 0
    image_rows_total: int = 0
    warnings: tuple[str, ...] = ()


@dataclass
class ExcelCatalogImportData:
    mapping: dict[str, tuple[str, str]]
    image_map: dict[str, list[str]]
    image_rows_total: int
    warnings: list[str] = field(default_factory=list)


OverwriteDecision = Callable[[str], bool]
ProgressCallback = Callable[[int, int, int, int], None]


def import_excel_catalog(
    xlsx_path: str | Path,
    *,
    db,
    assets_dir: str | Path,
    should_update_existing: OverwriteDecision,
    progress_callback: ProgressCallback | None = None,
) -> ExcelCatalogImportResult:
    data = read_excel_catalog_data(xlsx_path, assets_dir=assets_dir)
    return apply_excel_catalog_data(
        data,
        xlsx_path=xlsx_path,
        db=db,
        should_update_existing=should_update_existing,
        progress_callback=progress_callback,
    )


def apply_excel_catalog_data(
    data: ExcelCatalogImportData,
    *,
    xlsx_path: str | Path,
    db,
    should_update_existing: OverwriteDecision,
    progress_callback: ProgressCallback | None = None,
) -> ExcelCatalogImportResult:
    connection = db.connect()
    try:
        rows = connection.execute("SELECT code FROM items").fetchall()
        db_codes = [str(row["code"]) for row in rows]
        db_code_set = set(db_codes)
        db_code_index = build_unique_normalized_code_index(db_codes)
        automatically_approved: set[str] = set()

        updated = 0
        missing = 0
        missing_codes: list[str] = []
        images_updated = 0
        images_missing = 0
        skipped_existing = 0
        created_new = 0
        total = len(data.mapping)

        for index, (excel_code, descriptions) in enumerate(
            data.mapping.items(),
            start=1,
        ):
            code = str(excel_code).strip()
            matched_code = _match_database_code(code, db_code_set, db_code_index)
            if matched_code:
                if (
                    matched_code not in automatically_approved
                    and not should_update_existing(matched_code)
                ):
                    skipped_existing += 1
                    continue
                description_vi, description_en = descriptions
                cursor = connection.execute(
                    "UPDATE items SET description_excel=?, description_vietnames_from_excel=? WHERE code=?",
                    (
                        str(description_en).strip(),
                        str(description_vi).strip(),
                        matched_code,
                    ),
                )
                if cursor.rowcount > 0:
                    updated += 1
                else:
                    missing += 1
                missing_codes.append(code)
            else:
                description_vi, description_en = descriptions
                _insert_minimal_item(
                    connection,
                    code=code,
                    description_vi=str(description_vi).strip(),
                    description_en=str(description_en).strip(),
                )
                db_code_set.add(code)
                automatically_approved.add(code)
                created_new += 1

            if index % 25 == 0 and progress_callback is not None:
                progress_callback(index, total, updated, missing)

        excel_pdf_path = db.to_db_path(f"excel:{xlsx_path}")
        for excel_code, image_paths in data.image_map.items():
            unique_paths = list(dict.fromkeys(image_paths))
            code = str(excel_code).strip()
            matched_code = _match_database_code(code, db_code_set, db_code_index)
            if not matched_code:
                _insert_minimal_item(connection, code=code)
                db_code_set.add(code)
                automatically_approved.add(code)
                created_new += 1
                matched_code = code

            if (
                matched_code not in automatically_approved
                and not should_update_existing(matched_code)
            ):
                continue
            row = connection.execute(
                "SELECT id FROM items WHERE code=?",
                (matched_code,),
            ).fetchone()
            if row is None:
                images_missing += 1
                continue

            item_id = int(row["id"])
            connection.execute(
                "DELETE FROM item_asset_links WHERE item_id=?",
                (item_id,),
            )
            for image_index, image_path in enumerate(unique_paths):
                asset_path = db.to_db_path(image_path)
                asset_row = connection.execute(
                    "SELECT id FROM assets WHERE pdf_path=? AND page=? AND asset_path=?",
                    (excel_pdf_path, 0, asset_path),
                ).fetchone()
                if asset_row:
                    asset_id = int(asset_row["id"])
                else:
                    cursor = connection.execute(
                        """
                        INSERT INTO assets(pdf_path, page, asset_path, x0, y0, x1, y1, source, sha256)
                        VALUES(?,?,?,?,?,?,?,?,?)
                        """,
                        (
                            excel_pdf_path,
                            0,
                            asset_path,
                            None,
                            None,
                            None,
                            None,
                            "excel",
                            "",
                        ),
                    )
                    asset_id = int(cursor.lastrowid)
                connection.execute(
                    """
                    INSERT OR IGNORE INTO item_asset_links(item_id, asset_id, match_method, score, verified, is_primary)
                    VALUES(?,?,?,?,?,?)
                    """,
                    (
                        item_id,
                        asset_id,
                        "excel",
                        None,
                        1,
                        1 if image_index == 0 else 0,
                    ),
                )
            images_updated += 1

        connection.commit()
        return ExcelCatalogImportResult(
            total=total,
            created_new=created_new,
            updated=updated,
            skipped_existing=skipped_existing,
            missing=missing,
            missing_codes=tuple(missing_codes),
            images_updated=images_updated,
            images_missing=images_missing,
            image_rows_total=data.image_rows_total,
            warnings=tuple(data.warnings),
        )
    finally:
        connection.close()


def _match_database_code(
    code: str,
    exact_codes: set[str],
    normalized_index: dict[str, str],
) -> str:
    if code in exact_codes:
        return code
    return normalized_index.get(normalize_code_soft(code), "")


def _insert_minimal_item(
    connection,
    *,
    code: str,
    description_vi: str = "",
    description_en: str = "",
) -> None:
    connection.execute(
        """
        INSERT INTO items(
            code, description, description_excel,
            description_vietnames_from_excel, pdf_path, page,
            validated, validated_at, category, author, dimension,
            small_description, shape, blade_tip, surface_treatment, material
        )
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            code,
            "",
            description_en,
            description_vi,
            "",
            None,
            0,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ),
    )


def read_excel_catalog_data(
    xlsx_path: str | Path,
    *,
    assets_dir: str | Path,
) -> ExcelCatalogImportData:
    path = Path(xlsx_path)
    workbook = load_workbook(path)
    try:
        sheet_names = list(workbook.sheetnames)
        mapping = _read_description_mapping(path, sheet_names)
        image_map, image_rows_total, warnings = _extract_embedded_images(
            path,
            workbook,
            sheet_names,
            Path(assets_dir) / "excel_import",
        )
        return ExcelCatalogImportData(
            mapping=mapping,
            image_map=image_map,
            image_rows_total=image_rows_total,
            warnings=warnings,
        )
    finally:
        workbook.close()


def _read_description_mapping(
    xlsx_path: Path,
    sheet_names: list[str],
) -> dict[str, tuple[str, str]]:
    mapping: dict[str, tuple[str, str]] = {}
    for sheet_name in sheet_names:
        try:
            sheet_mapping = load_code_to_vi_en_from_excel(
                xlsx_path,
                sheet_name=sheet_name,
            )
        except Exception:
            continue
        for code, descriptions in sheet_mapping.items():
            key = str(code).strip()
            if key not in mapping:
                mapping[key] = (
                    str(descriptions[0]).strip(),
                    str(descriptions[1]).strip(),
                )
    return mapping


def _find_code_column_index(worksheet, header_row: int, code_column: str):
    normalized_code_column = normalize_header_text(code_column)
    for cell in worksheet[header_row]:
        if normalize_header_text(str(cell.value or "")) == normalized_code_column:
            return cell.column
    return None


def _read_code_rows(worksheet, header_row: int, code_column_index: int):
    code_rows: list[int] = []
    row_to_code: dict[int, str] = {}
    for row_number in range(header_row + 1, worksheet.max_row + 1):
        value = worksheet.cell(row=row_number, column=code_column_index).value
        code = str(value or "").strip()
        if code:
            code_rows.append(row_number)
            row_to_code[row_number] = code
    return code_rows, row_to_code


def _extract_embedded_images(
    xlsx_path: Path,
    workbook,
    sheet_names: list[str],
    output_dir: Path,
) -> tuple[dict[str, list[str]], int, list[str]]:
    warnings: list[str] = []
    image_map: dict[str, list[str]] = {}
    image_rows_total = 0
    hash_to_path: dict[str, str] = {}
    per_code_hashes: dict[str, set[str]] = {}
    first_code_occurrence: dict[str, tuple[str, int]] = {}
    sheet_code_data: dict[str, tuple[list[int], dict[int, str]]] = {}

    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        try:
            _frame, header_row, code_column = detect_excel_code_column(
                xlsx_path,
                sheet_name=sheet_name,
            )
        except Exception:
            continue
        header_row += 1
        code_column_index = _find_code_column_index(
            worksheet,
            header_row,
            code_column,
        )
        if code_column_index is None:
            continue
        code_rows, row_to_code = _read_code_rows(
            worksheet,
            header_row,
            code_column_index,
        )
        sheet_code_data[sheet_name] = (code_rows, row_to_code)
        for row_number in code_rows:
            code = row_to_code[row_number]
            if code not in first_code_occurrence:
                first_code_occurrence[code] = (sheet_name, row_number)

    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        code_rows, row_to_code = sheet_code_data.get(sheet_name, ([], {}))
        images = getattr(worksheet, "_images", None)
        if not code_rows or not images:
            continue
        output_dir.mkdir(parents=True, exist_ok=True)
        safe_sheet = sanitize_filename(sheet_name)

        for image in images:
            anchor_row = get_image_anchor_row(image)
            if anchor_row is None:
                continue
            image_rows_total += 1
            index = bisect_right(code_rows, anchor_row) - 1
            if index < 0:
                continue
            code_row = code_rows[index]
            code = row_to_code.get(code_row, "")
            if not code or first_code_occurrence.get(code) != (sheet_name, code_row):
                continue
            pil_image = image_to_pil(image)
            if pil_image is None:
                continue
            try:
                buffer = io.BytesIO()
                pil_image.convert("RGBA").save(buffer, format="PNG")
                data = buffer.getvalue()
            finally:
                pil_image.close()
            image_hash = sha256_bytes(data)
            code_hashes = per_code_hashes.setdefault(code, set())
            if image_hash in code_hashes:
                continue
            code_hashes.add(image_hash)

            output_path = output_dir / (
                f"{safe_sheet}_{sanitize_filename(code)}_{image_hash[:10]}.png"
            )
            path_string = hash_to_path.get(image_hash)
            if path_string is None:
                try:
                    output_path.write_bytes(data)
                except Exception as exc:
                    warnings.append(f"{sheet_name}:{code}: {exc}")
                    continue
                path_string = str(output_path)
                hash_to_path[image_hash] = path_string
            paths = image_map.setdefault(code, [])
            if path_string not in paths:
                paths.append(path_string)

    return image_map, image_rows_total, warnings
