from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from smartcatalog.loader.excel_loader import detect_excel_code_column
from smartcatalog.utils.text_normalization import normalize_header_text


@dataclass(frozen=True)
class WorkbookProductRow:
    code: str
    quantity: str


def read_workbook_products(xlsx_path: str | Path) -> tuple[WorkbookProductRow, ...]:
    _frame, header_row, code_column = detect_excel_code_column(xlsx_path)
    workbook = load_workbook(xlsx_path, read_only=True)
    try:
        worksheet = workbook.worksheets[0]
        header_row += 1
        code_column_index = _find_header_column(
            worksheet,
            header_row,
            code_column,
        )
        if code_column_index is None:
            raise ValueError(
                f"Không tìm thấy cột mã '{code_column}' trong dòng tiêu đề Excel."
            )

        quantity_column_index = _find_quantity_column(worksheet, header_row)
        rows: list[WorkbookProductRow] = []
        for row_number in range(header_row + 1, worksheet.max_row + 1):
            raw_code = worksheet.cell(
                row=row_number,
                column=code_column_index,
            ).value
            code = str(raw_code or "").strip()
            if not code:
                continue
            quantity = ""
            if quantity_column_index is not None:
                raw_quantity = worksheet.cell(
                    row=row_number,
                    column=quantity_column_index,
                ).value
                quantity = "" if raw_quantity is None else str(raw_quantity).strip()
            rows.append(WorkbookProductRow(code=code, quantity=quantity))
        return tuple(rows)
    finally:
        workbook.close()


def _find_header_column(worksheet, header_row: int, header: str):
    normalized_header = normalize_header_text(header)
    for cell in worksheet[header_row]:
        if normalize_header_text(str(cell.value or "")) == normalized_header:
            return cell.column
    return None


def _find_quantity_column(worksheet, header_row: int):
    quantity_headers = {"qty", "quantity", "so luong", "số lượng", "sl"}
    for cell in worksheet[header_row]:
        header = normalize_header_text(str(cell.value or ""))
        if header in quantity_headers or header.startswith("qty"):
            return cell.column
    return None
