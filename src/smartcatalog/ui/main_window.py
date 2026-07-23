# smartcatalog/ui/main_window.py
from __future__ import annotations

import threading
import io
import traceback
import tkinter as tk
import tkinter.font as tkfont
from tkinter import ttk, scrolledtext, messagebox, filedialog
from pathlib import Path
from typing import Callable, Optional, TYPE_CHECKING
import sqlite3
import os
import shutil
import datetime
if TYPE_CHECKING:
    from PIL import Image, ImageTk
from smartcatalog.state import AppState, CatalogItem
from smartcatalog.loader.pdf_loader import build_or_update_db_from_pdf
from smartcatalog.ui.widgets.scrollable_frame import ScrollableFrame
from smartcatalog.ui.controllers.candidates_controller import CandidatesControllerMixin
from smartcatalog.ui.controllers.images_controller import ImagesControllerMixin
from smartcatalog.ui.controllers.items_controller import ItemsControllerMixin
from smartcatalog.ui.controllers.item_form_controller import ItemFormControllerMixin
from smartcatalog.loader.excel_loader import load_code_to_vi_en_from_excel, detect_excel_code_column
from smartcatalog.ui.pdf_crop_window import PdfCropWindow
from smartcatalog.ui.export_review_dialog import ExportPreflightItem, ExportReviewDialog
from smartcatalog.utils.description_dictionary import import_dictionary_into_db
from smartcatalog.utils.post_processing import (
    DEFAULT_SHEET_NAME as POST_PROCESSING_SHEET_NAME,
    PostProcessingRow,
    apply_post_processing_branding,
    write_post_processing_sheet,
)

import re
import hashlib
from bisect import bisect_right
def _normalize_code_soft(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"\s+", "", s)  # remove all spaces
    return s

def _normalize_header_text(s: str) -> str:
    s = str(s or "").strip().lower()
    s = s.replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _sanitize_filename(s: str) -> str:
    s = re.sub(r"[^A-Za-z0-9_-]+", "_", str(s or "").strip())
    s = s.strip("_")
    return s or "item"


def _get_image_anchor_row(img) -> Optional[int]:
    try:
        anchor = getattr(img, "anchor", None)
        if anchor is None:
            return None
        if hasattr(anchor, "_from") and getattr(anchor._from, "row", None) is not None:
            return int(anchor._from.row) + 1
        if getattr(anchor, "row", None) is not None:
            return int(anchor.row) + 1
    except Exception:
        return None
    return None


def _image_to_pil(img) -> Optional[Image.Image]:
    from PIL import Image
    try:
        data = img._data()
        return Image.open(io.BytesIO(data))
    except Exception:
        pass
    try:
        ref = getattr(img, "ref", None)
        if ref:
            p = Path(ref)
            if p.exists():
                return Image.open(p)
    except Exception:
        return None
    return None

def _build_db_code_index(db_codes: list[str]) -> dict[str, str]:
    """
    normalized_code -> original_db_code
    only keep unique mappings to avoid wrong updates.
    """
    buckets: dict[str, list[str]] = {}
    for c in db_codes:
        key = _normalize_code_soft(c)
        buckets.setdefault(key, []).append(c)

    index: dict[str, str] = {}
    for k, vals in buckets.items():
        if len(vals) == 1:
            index[k] = vals[0]
    return index

def _safe_ui(root: tk.Misc, fn: Callable[[], None]) -> None:
    root.after(0, fn)


class MainWindow(
                    ttk.Frame,
                    ItemsControllerMixin,
                    CandidatesControllerMixin,
                    ImagesControllerMixin,
                    ItemFormControllerMixin,
                ):
    ...

    def __init__(self, root: tk.Tk, state: Optional[AppState] = None):
        super().__init__(root, padding=10)
        self.root = root
        self.state = state or AppState()
        self.state.ensure_dirs()

        self._sort_col: str = "id"
        self._sort_desc: bool = False

        self.status_message = tk.StringVar(value="Chưa tải dữ liệu")
        self._busy = tk.BooleanVar(value=False)

        # form vars
        self.var_code = tk.StringVar()
        self.var_page = tk.StringVar()

        self.var_category = tk.StringVar()
        self.var_author = tk.StringVar()
        self.var_dimension = tk.StringVar()
        self.var_small_description = tk.StringVar()
        self.var_shape = tk.StringVar()
        self.var_blade_tip = tk.StringVar()
        self.var_surface_treatment = tk.StringVar()
        self.var_material = tk.StringVar()
        self.var_validated = tk.BooleanVar(value=False)
        self.var_export_desc_en = tk.BooleanVar(value=False)
        self.var_export_desc_vi = tk.BooleanVar(value=False)

        self._thumb_refs: list[ImageTk.PhotoImage] = []
        self._full_img_ref: Optional[ImageTk.PhotoImage] = None
        self._selected_image_path: Optional[str] = None
        self.var_export_images_in_ui_order = tk.BooleanVar(value=False)

        self._selected: Optional[CatalogItem] = None
        self._column_drag_source_name: Optional[str] = None
        self._column_drag_start_x: Optional[int] = None
        self._column_drag_did_move: bool = False
        self._column_drag_clear_job: Optional[str] = None
        self._suppress_next_heading_sort: bool = False
        self._column_visibility_vars: dict[str, tk.BooleanVar] = {}
        self._hidden_column_positions: dict[str, int] = {}
        self._column_menu: Optional[tk.Menu] = None
        

        self._build_layout()

        self._build_left_panel()
        self._build_right_panel()
        self._update_pdf_tools_label()
        self._build_status_bar()

        self.root.after(0, self.refresh_items)

    # -----------------
    # Layout
    # -----------------

    def _build_layout(self) -> None:
        self.pack(fill="both", expand=True)
        self.root.title("SmartCatalog — Trình quản lý danh mục sản phẩm y tế")

        self.toolbar = ttk.Frame(self)
        self.toolbar.pack(fill="x", pady=(0, 8))

        self.btn_build_pdf = ttk.Button(self.toolbar, text="Tạo/Cập nhật CSDL từ PDF", command=self.on_choose_pdf_and_build_db)
        self.btn_build_pdf.pack(side="left", padx=(0, 6))
        
        self.btn_match_excel = ttk.Button(self.toolbar, text="Cập nhật CSDL từ Excel", command=self.on_build_excel_db)
        self.btn_match_excel.pack(side="left")

        self.btn_backup = ttk.Button(self.toolbar, text="💾 Backup CSDL", command=self.on_backup_data)
        # Hidden for now

        self.btn_refresh = ttk.Button(self.toolbar, text="🔄 Làm mới", command=self.refresh_items)

        ttk.Separator(self.toolbar, orient="vertical").pack(side="left", fill="y", padx=8)

        self.btn_search_images = ttk.Button(
            self.toolbar,
            text="Xuất catalog từ danh sách Excel",
            command=self.on_search_images_from_excel,
        )
        self.btn_search_images.pack(side="left")
        self.btn_search_images_opts = ttk.Button(
            self.toolbar,
            text="Tùy chọn xuất ▼",
            command=self._toggle_search_options_popup,
        )
        self._search_opts_popup: Optional[tk.Toplevel] = None
        self._search_opts_root_bind_id: Optional[str] = None

        # Panes
        self.panes = ttk.PanedWindow(self, orient="horizontal")
        self.panes.pack(side="top", fill="both", expand=True)

        self.left_pane = ttk.Frame(self.panes)
        self.right_pane = ttk.Frame(self.panes)

        self.panes.add(self.left_pane, weight=3)
        self.panes.add(self.right_pane, weight=2)

        # Scrollable container inside right pane
        self.right_scroll = ScrollableFrame(self.right_pane)
        self.right_scroll.pack(fill="both", expand=True)

        # Initial split: left 60%, right 40%
        self.root.after(80, self._set_initial_pane_ratio)

    def _set_initial_pane_ratio(self) -> None:
        try:
            total = int(self.panes.winfo_width() or 0)
            if total <= 1:
                self.root.after(80, self._set_initial_pane_ratio)
                return
            self.panes.sashpos(0, int(total * 0.6))
        except Exception:
            return


    def _build_left_panel(self) -> None:
        search_frame = ttk.Frame(self.left_pane)
        search_frame.pack(fill="x", pady=(0, 6))

        ttk.Label(search_frame, text="Tìm:").pack(side="left")
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(search_frame, textvariable=self.search_var)
        self.search_entry.pack(side="left", fill="x", expand=True, padx=6)
        self.search_entry.bind("<KeyRelease>", lambda _e: self._filter_items())

        list_frame = ttk.LabelFrame(self.left_pane, text="📦 Sản phẩm", padding=6)
        list_frame.pack(fill="both", expand=True)

        columns = (
            "id",
            "code",
            "page",
            "category",
            "author",
            "shape",
            "blade_tip",
            "dimension",
            "surface_treatment",
            "material",
            "small_description",
            "description_excel",
            "description_vietnames_from_excel",
            "validated",
        )
        self.items_tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=18)
        self.items_tree["displaycolumns"] = columns
        # Highlight matched rows when filtering
        base_font = tkfont.nametofont(ttk.Style().lookup("Treeview", "font") or "TkDefaultFont")
        match_font = base_font.copy()
        match_font.configure(weight="bold")
        self.items_tree.tag_configure("match", background="#fff4c2", font=match_font)
        self.items_tree.heading("id", command=lambda: self._sort_by("id"))
        self.items_tree.heading("code",command=lambda: self._sort_by("code"))
        self.items_tree.heading("page", command=lambda: self._sort_by("page"))
        self.items_tree.heading("category", command=lambda: self._sort_by("category"))
        self.items_tree.heading("author", command=lambda: self._sort_by("author"))
        self.items_tree.heading("shape", command=lambda: self._sort_by("shape"))
        self.items_tree.heading("blade_tip", command=lambda: self._sort_by("blade_tip"))
        self.items_tree.heading("dimension", command=lambda: self._sort_by("dimension"))
        self.items_tree.heading("surface_treatment", command=lambda: self._sort_by("surface_treatment"))
        self.items_tree.heading("material", command=lambda: self._sort_by("material"))
        self.items_tree.heading("small_description", command=lambda: self._sort_by("small_description"))
        self.items_tree.heading("description_excel", command=lambda: self._sort_by("description_excel"))
        self.items_tree.heading("description_vietnames_from_excel", command=lambda: self._sort_by("description_vietnames_from_excel"))
        self.items_tree.heading("validated", command=lambda: self._sort_by("validated"))

        self.items_tree.column("id", width=40, anchor="center", stretch=False)
        self.items_tree.column("code", width=90, anchor="w", stretch=False)
        self.items_tree.column("page", width=40, anchor="center", stretch=False)
        self.items_tree.column("category", width=100, anchor="w", stretch=False)
        self.items_tree.column("author", width=90, anchor="w", stretch=False)
        self.items_tree.column("shape", width=80, anchor="w", stretch=False)
        self.items_tree.column("blade_tip", width=80, anchor="w", stretch=False)
        self.items_tree.column("dimension", width=80, anchor="w", stretch=False)
        self.items_tree.column("surface_treatment", width=120, anchor="w", stretch=False)
        self.items_tree.column("material", width=80, anchor="w", stretch=False)
        self.items_tree.column("small_description", width=220, anchor="w", stretch=False)
        self.items_tree.column("description_excel", width=220, anchor="w", stretch=False)
        self.items_tree.column("description_vietnames_from_excel", width=220, anchor="w", stretch=False)
        self.items_tree.column("validated", width=150, anchor="center", stretch=False)

        yscroll = ttk.Scrollbar(list_frame, orient="vertical", command=self.items_tree.yview)
        xscroll = ttk.Scrollbar(list_frame, orient="horizontal", command=self.items_tree.xview)
        self.items_tree.configure(yscrollcommand=yscroll.set)
        self.items_tree.configure(xscrollcommand=xscroll.set)

        # Use grid to keep scrollbars aligned with the treeview
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)
        self.items_tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")

        self.items_tree.bind("<<TreeviewSelect>>", self._on_select_item)
        self.items_tree.bind("<ButtonPress-1>", self._on_items_tree_click_before_select, add="+")
        self.items_tree.bind("<ButtonPress-1>", self._on_items_tree_heading_press, add="+")
        self.items_tree.bind("<B1-Motion>", self._on_items_tree_heading_drag, add="+")
        self.items_tree.bind("<ButtonRelease-1>", self._on_items_tree_heading_release, add="+")
        self.items_tree.bind("<Button-3>", self._on_items_tree_heading_context_menu, add="+")
        self._update_sort_headers()

    def _get_display_columns(self) -> list[str]:
        cols = list(self.items_tree["displaycolumns"])
        if len(cols) == 1 and str(cols[0]) == "#all":
            cols = list(self.items_tree["columns"])
        return [str(c) for c in cols]

    def _on_items_tree_heading_press(self, event) -> None:
        if self.items_tree.identify_region(event.x, event.y) != "heading":
            self._column_drag_source_name = None
            self._column_drag_start_x = None
            self._column_drag_did_move = False
            return

        col_id = str(self.items_tree.identify_column(event.x) or "")
        if not col_id.startswith("#"):
            return
        try:
            idx = int(col_id[1:]) - 1
        except Exception:
            return

        cols = self._get_display_columns()
        if idx < 0 or idx >= len(cols):
            return

        self._column_drag_source_name = cols[idx]
        self._column_drag_start_x = int(event.x)
        self._column_drag_did_move = False

    def _on_items_tree_heading_drag(self, event) -> None:
        if not self._column_drag_source_name or self._column_drag_start_x is None:
            return

        if abs(int(event.x) - int(self._column_drag_start_x)) < 8:
            return

        col_id = str(self.items_tree.identify_column(event.x) or "")
        if not col_id.startswith("#"):
            return
        try:
            target_idx = int(col_id[1:]) - 1
        except Exception:
            return

        cols = self._get_display_columns()
        if self._column_drag_source_name not in cols:
            return

        src_idx = cols.index(self._column_drag_source_name)
        if target_idx < 0 or target_idx >= len(cols) or target_idx == src_idx:
            return

        moving = cols.pop(src_idx)
        cols.insert(target_idx, moving)
        self.items_tree["displaycolumns"] = tuple(cols)
        self._column_drag_did_move = True

    def _clear_heading_sort_suppress(self) -> None:
        self._column_drag_clear_job = None
        self._suppress_next_heading_sort = False
        self._column_drag_did_move = False

    def _on_items_tree_heading_release(self, _event) -> None:
        moved = bool(self._column_drag_did_move)
        self._column_drag_source_name = None
        self._column_drag_start_x = None

        if moved:
            self._suppress_next_heading_sort = True
            if self._column_drag_clear_job is not None:
                try:
                    self.root.after_cancel(self._column_drag_clear_job)
                except Exception:
                    pass
            self._column_drag_clear_job = self.root.after(250, self._clear_heading_sort_suppress)
        else:
            self._column_drag_did_move = False

    @staticmethod
    def _column_label(col: str) -> str:
        labels = {
            "id": "No",
            "code": "Mã",
            "page": "Trang",
            "category": "Chủng loại",
            "author": "Tác giả",
            "shape": "Hình dạng",
            "blade_tip": "Đầu lưỡi",
            "dimension": "Kích thước",
            "surface_treatment": "Xử lý bề mặt/ công nghệ",
            "material": "Material",
            "small_description": "Mô tả từ PDF",
            "description_excel": "Mô tả EN từ Excel",
            "description_vietnames_from_excel": "Mô tả VI từ Excel",
            "validated": "Đã kiểm duyệt (Thời gian)",
        }
        return labels.get(str(col), str(col))

    def _sync_column_visibility_vars(self) -> None:
        visible = set(self._get_display_columns())
        for col in [str(c) for c in self.items_tree["columns"]]:
            var = self._column_visibility_vars.get(col)
            if var is None:
                var = tk.BooleanVar(value=(col in visible))
                self._column_visibility_vars[col] = var
            else:
                var.set(col in visible)

    def _toggle_column_visibility(self, col: str) -> None:
        col = str(col)
        visible = self._get_display_columns()
        if col not in [str(c) for c in self.items_tree["columns"]]:
            return

        var = self._column_visibility_vars.get(col)
        wants_visible = bool(var.get()) if var is not None else (col not in visible)

        if wants_visible:
            if col in visible:
                return
            insert_at = self._hidden_column_positions.pop(col, len(visible))
            insert_at = max(0, min(int(insert_at), len(visible)))
            visible.insert(insert_at, col)
            self.items_tree["displaycolumns"] = tuple(visible)
            self._update_sort_headers()
            self._sync_column_visibility_vars()
            return

        if col not in visible:
            return
        if len(visible) <= 1:
            if var is not None:
                var.set(True)
            messagebox.showwarning("Không thể ẩn", "Phải giữ lại ít nhất một cột hiển thị.")
            return

        self._hidden_column_positions[col] = visible.index(col)
        visible = [c for c in visible if c != col]
        self.items_tree["displaycolumns"] = tuple(visible)
        self._update_sort_headers()
        self._sync_column_visibility_vars()

    def _on_items_tree_heading_context_menu(self, event) -> str | None:
        if self.items_tree.identify_region(event.x, event.y) != "heading":
            return None

        try:
            if self._column_menu is not None:
                self._column_menu.destroy()
        except Exception:
            pass

        self._sync_column_visibility_vars()
        menu = tk.Menu(self.items_tree, tearoff=0)
        for col in [str(c) for c in self.items_tree["columns"]]:
            var = self._column_visibility_vars[col]
            menu.add_checkbutton(
                label=self._column_label(col),
                variable=var,
                command=lambda c=col: self._toggle_column_visibility(c),
            )

        self._column_menu = menu
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            try:
                menu.grab_release()
            except Exception:
                pass
        return "break"

    def _build_right_panel(self) -> None:
        parent = self.right_scroll.inner

        self._build_item_editor_section(parent)
        self._build_candidates_section_simple(parent)


    def _build_item_editor_section(self, parent) -> None:
        # "Thêm mới" sits above the "Sản phẩm" section.
        add_row = ttk.Frame(parent)
        add_row.pack(fill="x", pady=(0, 4))
        add_row.columnconfigure(0, weight=1)
        self.btn_add_item = ttk.Button(add_row, text="➕ Thêm mới", command=self.on_add_item)
        self.btn_add_item.grid(row=0, column=1, sticky="e")

        editor = ttk.LabelFrame(parent, text="🧾 Sản phẩm", padding=8)
        editor.pack(fill="x", pady=(0, 0))
        editor.columnconfigure(1, weight=1)

        # --- Action row: PDF info + Save/Delete ---
        top = ttk.Frame(editor)
        top.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 8))
        top.columnconfigure(0, weight=1)

        self.pdf_tools_label = ttk.Label(top, text="Chưa chọn PDF")
        self.pdf_tools_label.grid(row=0, column=0, sticky="w")
        self.btn_save = ttk.Button(top, text="💾 Lưu", command=self.on_save_item)
        self.btn_save.grid(row=0, column=1, sticky="e", padx=(0, 6))
        self.btn_delete_item = ttk.Button(top, text="🗑️ Xóa", command=self.on_delete_item)
        self.btn_delete_item.grid(row=0, column=2, sticky="e")
        self.chk_validated = ttk.Checkbutton(top, text="Đã kiểm duyệt", variable=self.var_validated)
        self.chk_validated.grid(row=1, column=1, columnspan=2, sticky="e", pady=(4, 0))

        # --- Fields (replaces "Item fields" box) ---
        r = 1

        ttk.Label(editor, text="Mã").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_code).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="Trang").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_page).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="Chủng loại").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_category).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="Tác giả").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_author).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="Hình dạng").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_shape).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="Đầu lưỡi").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_blade_tip).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="Kích thước").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_dimension).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="Xử lý bề mặt/ công nghệ").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_surface_treatment).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="Material").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_material).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="Mô tả từ PDF").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        ttk.Entry(editor, textvariable=self.var_small_description).grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="Mô tả EN từ Excel").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        self.description_excel_text = scrolledtext.ScrolledText(editor, wrap="word", height=4)
        self.description_excel_text.grid(row=r, column=1, sticky="ew", pady=3)
        r += 1

        ttk.Label(editor, text="Mô tả VI từ Excel").grid(row=r, column=0, sticky="w", padx=(0, 8), pady=3)
        self.description_vietnames_from_excel_text = scrolledtext.ScrolledText(editor, wrap="word", height=4)
        self.description_vietnames_from_excel_text.grid(row=r, column=1, sticky="ew", pady=3)

        # Ảnh là một phần của nhóm "Sản phẩm"
        images_holder = ttk.Frame(editor)
        images_holder.grid(row=r + 1, column=0, columnspan=2, sticky="ew", pady=(8, 0))
        self._build_images_section(images_holder)

    def on_open_pdf_cropper(self) -> None:
        if not self.state.catalog_pdf_path:
            messagebox.showwarning("Chưa có PDF", "Vui lòng tạo/chọn PDF trước.")
            return
        if not self._selected or not getattr(self._selected, "page", None):
            messagebox.showwarning("Thiếu trang", "Vui lòng chọn sản phẩm có trang hợp lệ.")
            return

        def after_save():
            self.refresh_items()
            self._selected = next((x for x in self.state.items_cache if x.id == self._selected.id), self._selected)
            self._reload_selected_into_form()

            if self._selected and getattr(self._selected, "page", None):
                self._render_candidates_for_page(int(self._selected.page) - 1)

            self._set_status("✅ Đã lưu cắt ảnh và gắn vào sản phẩm")


        PdfCropWindow(
            self.root,
            state=self.state,
            item_id=int(self._selected.id),
            page_1based=int(self._selected.page),
            on_after_save=after_save,
            title="Cắt từ PDF",
        )


    def _update_pdf_tools_label(self) -> None:
        pdf = self.state.catalog_pdf_path
        it = self._selected

        # If label not created yet (defensive)
        if not hasattr(self, "pdf_tools_label"):
            return

        if not pdf:
            self.pdf_tools_label.configure(text="Chưa chọn PDF")
            return

        pdf_name = Path(pdf).name if not isinstance(pdf, Path) else pdf.name
        page = getattr(it, "page", None) if it else None

        if page:
            self.pdf_tools_label.configure(text=f"PDF: {pdf_name} | Trang: {page}")
        else:
            self.pdf_tools_label.configure(text=f"PDF: {pdf_name} | (chọn sản phẩm)")


    def _build_images_section(self, parent) -> None:
        images_frame = ttk.LabelFrame(parent, text="🖼 Ảnh", padding=8)
        images_frame.pack(fill="both", expand=False, pady=(8, 0))

        content = ttk.Frame(images_frame)
        content.pack(fill="both", expand=True)

        btn_col = ttk.Frame(content)
        btn_col.pack(side="right", fill="y", padx=(8, 0))
        ttk.Button(btn_col, text="Thêm ảnh", command=self.on_add_image).pack(fill="x")
        ttk.Button(btn_col, text="Xoay 90°", command=lambda: self.on_rotate_selected_image(90)).pack(fill="x", pady=(4, 0))
        ttk.Button(btn_col, text="Xóa ảnh đã chọn", command=self.on_remove_selected_thumbnail).pack(fill="x", pady=(4, 0))

        thumb_host = ttk.Frame(content)
        thumb_host.pack(side="left", fill="both", expand=True)

        self.thumb_canvas = tk.Canvas(thumb_host, height=180, highlightthickness=0)
        thumb_scroll = ttk.Scrollbar(thumb_host, orient="vertical", command=self.thumb_canvas.yview)
        self.thumb_canvas.configure(yscrollcommand=thumb_scroll.set)

        thumb_scroll.pack(side="right", fill="y")
        self.thumb_canvas.pack(side="left", fill="both", expand=True)

        self.thumb_inner = ttk.Frame(self.thumb_canvas)
        self.thumb_canvas_window = self.thumb_canvas.create_window((0, 0), window=self.thumb_inner, anchor="nw")

        def _on_thumb_inner_configure(_e=None):
            self.thumb_canvas.configure(scrollregion=self.thumb_canvas.bbox("all"))

        def _on_thumb_canvas_configure(evt):
            self.thumb_canvas.itemconfig(self.thumb_canvas_window, width=evt.width)

        self.thumb_inner.bind("<Configure>", _on_thumb_inner_configure)
        self.thumb_canvas.bind("<Configure>", _on_thumb_canvas_configure)

    def _build_actions_section(self, parent) -> None:
        actions = ttk.Frame(parent)
        actions.pack(fill="x", pady=(8, 0))

        self.btn_save = ttk.Button(actions, text="💾 Lưu", command=self.on_save_item)
        self.btn_save.pack(side="left", padx=(0, 6))

        self.btn_reload = None
        self.btn_clear = None
    
    #-----------------------------------------------------


    def _build_status_bar(self) -> None:
        bar = ttk.Frame(self)
        bar.pack(side="bottom", fill="x", pady=(8, 0))

        self.progress = ttk.Progressbar(bar, mode="indeterminate")
        self.progress.pack(side="left", fill="x", expand=True, padx=(0, 8))

        self.status_bar = ttk.Label(bar, textvariable=self.status_message, anchor="w")
        self.status_bar.pack(side="left")

        self._apply_busy(False)

    # -----------------
    # Busy / status
    # -----------------

    def _apply_busy(self, busy: bool) -> None:
        self._busy.set(busy)
        for w in (self.btn_build_pdf, self.btn_refresh, self.btn_match_excel, self.btn_search_images,
                  self.btn_search_images_opts,
                  self.btn_save, self.btn_add_item, self.btn_delete_item, self.btn_backup):
            w.configure(state=("disabled" if busy else "normal"))
        if busy:
            self._close_search_options_popup()

        if busy:
            self.progress.start(10)
        else:
            self.progress.stop()

    def _set_status(self, msg: str) -> None:
        self.status_message.set(msg)

    def _set_preview_text(self, text: str) -> None:
        """
        Legacy no-op to keep controller calls safe after removing preview widget.
        """
        return

    def _toggle_search_options_popup(self) -> None:
        self._open_search_options_popup()

    def _open_search_options_popup(self) -> Optional[bool]:
        if self._search_opts_popup and self._search_opts_popup.winfo_exists():
            return None

        popup = tk.Toplevel(self.root)
        popup.transient(self.root)
        popup.title("Tùy chọn xuất")
        popup.resizable(False, False)

        frame = ttk.Frame(popup, padding=12)
        frame.pack(fill="both", expand=True)

        ttk.Checkbutton(
            frame,
            text="Xuất mô tả EN",
            variable=self.var_export_desc_en,
        ).pack(anchor="w")
        ttk.Checkbutton(
            frame,
            text="Xuất mô tả VI",
            variable=self.var_export_desc_vi,
        ).pack(anchor="w", pady=(4, 0))
        ttk.Checkbutton(
            frame,
            text="Giữ đúng thứ tự ảnh trong CSDL",
            variable=self.var_export_images_in_ui_order,
        ).pack(anchor="w", pady=(4, 0))

        btn_row = ttk.Frame(frame)
        btn_row.pack(fill="x", pady=(10, 0))
        btn_row.columnconfigure(0, weight=1)

        result = {"ok": False}

        def _on_ok() -> None:
            result["ok"] = True
            self._close_search_options_popup()

        def _on_cancel() -> None:
            result["ok"] = False
            self._close_search_options_popup()

        ttk.Button(btn_row, text="OK", command=_on_ok).grid(row=0, column=1, sticky="e")
        ttk.Button(btn_row, text="Hủy", command=_on_cancel).grid(row=0, column=2, sticky="e", padx=(6, 0))

        popup.protocol("WM_DELETE_WINDOW", _on_cancel)

        self._search_opts_popup = popup
        popup.update_idletasks()
        try:
            # Center on root
            x = self.root.winfo_rootx() + int(self.root.winfo_width() / 2) - int(popup.winfo_width() / 2)
            y = self.root.winfo_rooty() + int(self.root.winfo_height() / 2) - int(popup.winfo_height() / 2)
            popup.geometry(f"+{max(0, x)}+{max(0, y)}")
        except Exception:
            pass
        popup.grab_set()
        popup.focus_force()
        self.root.wait_window(popup)
        return bool(result["ok"])

    def _close_search_options_popup(self) -> None:
        if self._search_opts_popup and self._search_opts_popup.winfo_exists():
            try:
                self._search_opts_popup.destroy()
            except Exception:
                pass
        self._search_opts_popup = None

    @staticmethod
    def _is_descendant_widget(widget: tk.Misc, ancestor: tk.Misc) -> bool:
        cur = widget
        while cur is not None:
            if cur == ancestor:
                return True
            cur = getattr(cur, "master", None)
        return False

    def _run_bg(self, title: str, work: Callable[[], None]) -> None:
        def runner():
            try:
                _safe_ui(self.root, lambda: (self._apply_busy(True), self._set_status(title)))
                work()
                _safe_ui(self.root, lambda: self._apply_busy(False))
            except Exception as exc:
                tb = traceback.format_exc()

                # ✅ capture strings now
                err_text = f"{exc}\n\n{tb}"

                _safe_ui(self.root, lambda: self._apply_busy(False))
                _safe_ui(self.root, lambda: self._set_status(f"❌ Lỗi: {exc}"))
                _safe_ui(self.root, lambda msg=err_text: messagebox.showerror("Lỗi", msg))

        threading.Thread(target=runner, daemon=True).start()

    def _show_missing_codes(self, missing_codes: list[str]) -> None:
        if not missing_codes:
            return

        popup = tk.Toplevel(self.root)
        popup.title("Các mã không tìm thấy")
        popup.transient(self.root)
        popup.resizable(True, True)

        frame = ttk.Frame(popup, padding=12)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Các mã không tìm thấy trong CSDL").pack(anchor="w")

        list_frame = ttk.Frame(frame)
        list_frame.pack(fill="both", expand=True, pady=(8, 0))
        list_frame.rowconfigure(0, weight=1)
        list_frame.columnconfigure(0, weight=1)

        lb = tk.Listbox(list_frame, height=12)
        sb = ttk.Scrollbar(list_frame, orient="vertical", command=lb.yview)
        lb.configure(yscrollcommand=sb.set)

        lb.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")

        for code in missing_codes:
            lb.insert("end", str(code))

        btn_row = ttk.Frame(frame)
        btn_row.pack(fill="x", pady=(10, 0))
        btn_row.columnconfigure(0, weight=1)
        def _copy_all() -> None:
            try:
                popup.clipboard_clear()
                popup.clipboard_append("\n".join([str(x) for x in missing_codes]))
            except Exception:
                pass

        ttk.Button(btn_row, text="Sao chép tất cả các mã thiếu", command=_copy_all).grid(row=0, column=1, sticky="e")
        ttk.Button(btn_row, text="Đóng", command=popup.destroy).grid(row=0, column=2, sticky="e", padx=(6, 0))

        popup.update_idletasks()
        try:
            x = self.root.winfo_rootx() + int(self.root.winfo_width() / 2) - int(popup.winfo_width() / 2)
            y = self.root.winfo_rooty() + int(self.root.winfo_height() / 2) - int(popup.winfo_height() / 2)
            popup.geometry(f"+{max(0, x)}+{max(0, y)}")
        except Exception:
            pass

    def _prompt_open_excel(self, xlsx_path: str) -> None:
        if not xlsx_path:
            return
        try:
            try:
                os.startfile(xlsx_path)
            except Exception as e:
                messagebox.showwarning("Không mở được file", f"Không thể mở file Excel:\n{e}")
        except Exception:
            return

    def _show_export_result(
        self,
        *,
        xlsx_path: str,
        matched: int,
        total: int,
        images_found: int,
        missing_vi: int,
        missing_en: int,
        missing_images: int,
        missing_codes: list[str],
    ) -> None:
        """Show one combined export summary and unknown-code list."""
        popup = tk.Toplevel(self.root)
        popup.title("Xuất dữ liệu hoàn tất")
        popup.transient(self.root)
        popup.geometry("620x560")
        popup.minsize(520, 440)
        popup.grab_set()

        frame = ttk.Frame(popup, padding=14)
        frame.pack(fill="both", expand=True)

        ttk.Label(
            frame,
            text="Xuất catalog hoàn tất",
            font=("Segoe UI", 14, "bold"),
        ).pack(anchor="w")
        ttk.Label(
            frame,
            text=str(xlsx_path),
            foreground="#555555",
            wraplength=570,
        ).pack(anchor="w", pady=(3, 12))

        summary = ttk.LabelFrame(frame, text="Kết quả", padding=10)
        summary.pack(fill="x")
        summary_text = (
            f"Sản phẩm khớp: {matched}/{total}\n"
            f"Tìm thấy hình ảnh: {images_found}/{total}\n"
            f"Thiếu mô tả VI: {missing_vi}\n"
            f"Thiếu mô tả EN: {missing_en}\n"
            f"Thiếu hình ảnh: {missing_images}\n"
            f"Mã không tồn tại: {len(missing_codes)}"
        )
        ttk.Label(summary, text=summary_text, justify="left").pack(anchor="w")

        missing_frame = ttk.LabelFrame(
            frame,
            text=f"Các mã không tìm thấy ({len(missing_codes)})",
            padding=8,
        )
        missing_frame.pack(fill="both", expand=True, pady=(12, 0))
        missing_frame.rowconfigure(0, weight=1)
        missing_frame.columnconfigure(0, weight=1)

        if missing_codes:
            codes_list = tk.Listbox(missing_frame, height=9)
            codes_scroll = ttk.Scrollbar(
                missing_frame,
                orient="vertical",
                command=codes_list.yview,
            )
            codes_list.configure(yscrollcommand=codes_scroll.set)
            codes_list.grid(row=0, column=0, sticky="nsew")
            codes_scroll.grid(row=0, column=1, sticky="ns")
            for code in missing_codes:
                codes_list.insert("end", str(code))
        else:
            ttk.Label(
                missing_frame,
                text="Tất cả mã sản phẩm đều được tìm thấy trong CSDL.",
                foreground="#267326",
            ).grid(row=0, column=0, sticky="nw")

        buttons = ttk.Frame(frame)
        buttons.pack(fill="x", pady=(12, 0))

        def copy_missing_codes() -> None:
            if not missing_codes:
                return
            try:
                popup.clipboard_clear()
                popup.clipboard_append("\n".join(str(code) for code in missing_codes))
                popup.update()
            except Exception:
                pass

        copy_button = ttk.Button(
            buttons,
            text="Sao chép mã không tìm thấy",
            command=copy_missing_codes,
            state=("normal" if missing_codes else "disabled"),
        )
        copy_button.pack(side="left")
        ttk.Button(
            buttons,
            text="Mở file Excel",
            command=lambda: self._prompt_open_excel(xlsx_path),
        ).pack(side="right")
        ttk.Button(
            buttons,
            text="Đóng",
            command=popup.destroy,
        ).pack(side="right", padx=(0, 8))

        popup.protocol("WM_DELETE_WINDOW", popup.destroy)
        popup.update_idletasks()
        try:
            x = self.root.winfo_rootx() + (self.root.winfo_width() - popup.winfo_width()) // 2
            y = self.root.winfo_rooty() + (self.root.winfo_height() - popup.winfo_height()) // 2
            popup.geometry(f"+{max(0, x)}+{max(0, y)}")
        except Exception:
            pass

    # -----------------
    # Actions
    # -----------------

    def on_choose_pdf_and_build_db(self) -> None:
        """
        Show existing catalog PDFs, ask whether to load a new one, then build/update DB.
        """
        # show existing catalog PDFs
        pdf_dir = self.state.data_dir / "catalog_pdfs"
        existing = []
        try:
            if pdf_dir.exists():
                existing = sorted([p.name for p in pdf_dir.glob("*.pdf")])
            if existing:
                messagebox.showinfo(
                    "PDF danh mục đã có",
                    "Đã có trong CSDL:\n" + "\n".join(existing),
                )
        except Exception:
            pass

        use_new = messagebox.askyesno(
            "Tải PDF mới?",
            "Bạn có muốn tải một file PDF mới không?",
        )

        if use_new or not self.state.catalog_pdf_path:
            path = filedialog.askopenfilename(
                title="Chọn PDF danh mục",
                initialdir=str(pdf_dir) if pdf_dir.exists() else None,
                filetypes=[("Tệp PDF", "*.pdf"), ("Tất cả tệp", "*.*")],
            )
            if not path:
                return
            self.state.set_catalog_pdf(path)
            _safe_ui(self.root, self._update_pdf_tools_label)
            self._set_status(f"Đã chọn PDF: {path}")
        else:
            path = str(self.state.catalog_pdf_path)
            if not path:
                return
            run_again = messagebox.askyesno(
                "Cập nhật lại CSDL?",
                "Dùng PDF hiện tại và cập nhật lại?",
            )
            if not run_again:
                self._set_status("Đã hủy cập nhật PDF.")
                return
            self._set_status(f"Đang dùng PDF hiện tại: {path}")

        # From here: we have a PDF path
        def work():
            apply_all_choice: Optional[bool] = None

            def on_existing_item_decision(code: str) -> bool:
                nonlocal apply_all_choice
                if apply_all_choice is not None:
                    return bool(apply_all_choice)

                done = threading.Event()
                result: dict[str, bool] = {
                    "update": False,
                    "apply_all": False,
                }

                def ask_on_ui_thread():
                    update, apply_all = self._ask_update_existing_item_dialog(code)
                    result["update"] = bool(update)
                    result["apply_all"] = bool(apply_all)
                    done.set()

                _safe_ui(self.root, ask_on_ui_thread)
                done.wait()
                if result["apply_all"]:
                    apply_all_choice = result["update"]
                return result["update"]

            build_or_update_db_from_pdf(
                self.state,
                None,
                self.status_message,
                on_existing_item_decision=on_existing_item_decision,
            )
            _safe_ui(self.root, self.refresh_items)
            _safe_ui(self.root, lambda: self._set_status("✅ Cập nhật CSDL từ PDF xong"))

        self._run_bg("⏳ Đang tạo/cập nhật CSDL từ PDF...", work)

    def _ask_update_existing_item_dialog(self, code: str) -> tuple[bool, bool]:
        """
        Ask what to do when an item code already exists.
        Returns: (update_existing, apply_for_all_existing)
        """
        dlg = tk.Toplevel(self.root)
        dlg.title("Mã đã tồn tại")
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(False, False)

        frame = ttk.Frame(dlg, padding=12)
        frame.pack(fill="both", expand=True)

        ttk.Label(
            frame,
            text=f"Mã sản phẩm '{code}' đã có trong CSDL.\nBạn muốn cập nhật/ghi đè dữ liệu hiện tại không?",
            justify="left",
        ).pack(anchor="w")

        apply_all_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            frame,
            text="Áp dụng cho tất cả mã đã tồn tại",
            variable=apply_all_var,
        ).pack(anchor="w", pady=(10, 0))

        result = {"update": False}

        btns = ttk.Frame(frame)
        btns.pack(fill="x", pady=(12, 0))

        def choose_update():
            result["update"] = True
            dlg.destroy()

        def choose_skip():
            result["update"] = False
            dlg.destroy()

        ttk.Button(btns, text="Cập nhật/Ghi đè", command=choose_update).pack(side="left")
        ttk.Button(btns, text="Giữ nguyên (bỏ qua)", command=choose_skip).pack(side="left", padx=(8, 0))

        dlg.protocol("WM_DELETE_WINDOW", choose_skip)
        dlg.wait_window()
        return bool(result["update"]), bool(apply_all_var.get())

    def on_backup_data(self) -> None:
        """
        Backup SQLite DB and assets folder to a user-chosen directory.
        """
        if not self.state or not self.state.db_path:
            messagebox.showwarning("Thiếu CSDL", "Không tìm thấy đường dẫn CSDL để backup.")
            return

        db_path = Path(self.state.db_path)
        if not db_path.exists():
            messagebox.showwarning("Thiếu CSDL", f"Không tìm thấy file DB: {db_path}")
            return

        dest_root = filedialog.askdirectory(
            title="Chọn thư mục lưu backup",
            mustexist=True,
        )
        if not dest_root:
            return

        ts = datetime.datetime.now().strftime("%Y-%m-%d_%H%M")
        backup_dir = Path(dest_root) / f"smartcatalog_backup_{ts}"

        def work():
            try:
                backup_dir.mkdir(parents=True, exist_ok=True)

                # 1) SQLite backup (safe even if DB is open)
                backup_db_path = backup_dir / "catalog.db"
                src_conn = sqlite3.connect(str(db_path))
                try:
                    dst_conn = sqlite3.connect(str(backup_db_path))
                    try:
                        src_conn.backup(dst_conn)
                    finally:
                        dst_conn.close()
                finally:
                    src_conn.close()

                # 2) Assets backup
                assets_src = Path(self.state.assets_dir)
                assets_dst = backup_dir / "assets"
                if assets_src.exists():
                    shutil.copytree(assets_src, assets_dst, dirs_exist_ok=True)

                # 3) Catalog PDFs backup
                pdfs_src = self.state.data_dir / "catalog_pdfs"
                pdfs_dst = backup_dir / "catalog_pdfs"
                if pdfs_src.exists():
                    shutil.copytree(pdfs_src, pdfs_dst, dirs_exist_ok=True)

                # 4) Settings backup
                settings_src = Path(self.state.settings_path)
                settings_dst = backup_dir / "settings.json"
                if settings_src.exists():
                    shutil.copy2(settings_src, settings_dst)

                _safe_ui(
                    self.root,
                    lambda: messagebox.showinfo(
                        "Backup xong",
                        f"Đã backup CSDL và assets vào:\n{backup_dir}",
                    ),
                )
            except Exception as exc:
                _safe_ui(
                    self.root,
                    lambda: messagebox.showerror(
                        "Backup lỗi",
                        f"Không thể backup: {exc}",
                    ),
                )

        self._run_bg("⏳ Đang backup CSDL...", work)

    def on_build_excel_db(self) -> None:
        """
        Load an Excel file and update items.description_excel and images by matching item code.
        Matching strategy:
        1) exact code match
        2) normalized match (spaces removed, weird dashes fixed) -> only if uniquely maps to a DB code
        """
        if not self.state.db:
            messagebox.showwarning("Thiếu CSDL", "Vui lòng tạo/tải CSDL trước (từ PDF).")
            return

        xlsx_path = filedialog.askopenfilename(
            title="Chọn file Excel",
            filetypes=[("Tệp Excel", "*.xlsx *.xls"), ("Tất cả tệp", "*.*")],
        )
        if not xlsx_path:
            return

        def work():
            from openpyxl import load_workbook
            apply_all_choice: Optional[bool] = None
            decision_by_code: dict[str, bool] = {}

            def should_update_existing(code: str) -> bool:
                nonlocal apply_all_choice
                code_key = str(code or "").strip()
                if not code_key:
                    return False
                if code_key in decision_by_code:
                    return bool(decision_by_code[code_key])
                if apply_all_choice is not None:
                    decision_by_code[code_key] = bool(apply_all_choice)
                    return bool(apply_all_choice)

                done = threading.Event()
                result: dict[str, bool] = {"update": False, "apply_all": False}

                def ask_on_ui_thread():
                    update, apply_all = self._ask_update_existing_item_dialog(code_key)
                    result["update"] = bool(update)
                    result["apply_all"] = bool(apply_all)
                    done.set()

                _safe_ui(self.root, ask_on_ui_thread)
                done.wait()
                if result["apply_all"]:
                    apply_all_choice = result["update"]
                decision_by_code[code_key] = result["update"]
                return bool(result["update"])

            # 1) read excel (all sheets) -> {excel_code: (vi, en)}
            mapping: dict[str, tuple[str, str]] = {}
            wb = load_workbook(xlsx_path)
            sheet_names = list(wb.sheetnames)
            for sn in sheet_names:
                try:
                    m = load_code_to_vi_en_from_excel(xlsx_path, sheet_name=sn)
                except Exception:
                    continue
                for k, v in m.items():
                    key = str(k).strip()
                    if key in mapping:
                        continue
                    mapping[key] = (str(v[0]).strip(), str(v[1]).strip())

            # 2) read all DB codes once (exact + normalized index)
            conn = self.state.db.connect()
            try:
                rows = conn.execute("SELECT code FROM items").fetchall()
                db_codes = [str(r["code"]) for r in rows]
            finally:
                conn.close()

            db_code_set = set(db_codes)
            db_index = _build_db_code_index(db_codes)  # normalized -> original db code (unique only)

            # 3) build image map from Excel (embedded images)
            image_map: dict[str, list[str]] = {}
            image_rows_total = 0
            try:
                hash_to_path: dict[str, str] = {}
                per_code_hashes: dict[str, set[str]] = {}
                first_code_occurrence: dict[str, tuple[str, int]] = {}

                # Pre-pass: find first (sheet, row) occurrence of each code across all sheets
                for sn in sheet_names:
                    ws = wb[sn]
                    try:
                        _df, header_row, code_col = detect_excel_code_column(xlsx_path, sheet_name=sn)
                    except Exception:
                        continue
                    header_row_1 = header_row + 1  # openpyxl is 1-based

                    code_col_idx = None
                    for cell in ws[header_row_1]:
                        if _normalize_header_text(str(cell.value or "")) == _normalize_header_text(code_col):
                            code_col_idx = cell.column
                            break
                    if code_col_idx is None:
                        continue

                    for r in range(header_row_1 + 1, ws.max_row + 1):
                        raw_code = ws.cell(row=r, column=code_col_idx).value
                        excel_code_str = str(raw_code or "").strip()
                        if not excel_code_str:
                            continue
                        if excel_code_str not in first_code_occurrence:
                            first_code_occurrence[excel_code_str] = (sn, r)

                for sn in sheet_names:
                    ws = wb[sn]
                    try:
                        _df, header_row, code_col = detect_excel_code_column(xlsx_path, sheet_name=sn)
                    except Exception:
                        continue
                    header_row_1 = header_row + 1  # openpyxl is 1-based

                    # find code column index in header row
                    code_col_idx = None
                    for cell in ws[header_row_1]:
                        if _normalize_header_text(str(cell.value or "")) == _normalize_header_text(code_col):
                            code_col_idx = cell.column
                            break
                    if code_col_idx is None:
                        continue

                    # map row -> excel code
                    code_rows: list[int] = []
                    row_to_code: dict[int, str] = {}
                    for r in range(header_row_1 + 1, ws.max_row + 1):
                        raw_code = ws.cell(row=r, column=code_col_idx).value
                        excel_code_str = str(raw_code or "").strip()
                        if not excel_code_str:
                            continue
                        code_rows.append(r)
                        row_to_code[r] = excel_code_str
                    code_rows.sort()

                    # extract images and map to nearest code row above
                    if code_rows and getattr(ws, "_images", None):
                        out_dir = Path(self.state.assets_dir) / "excel_import"
                        out_dir.mkdir(parents=True, exist_ok=True)

                        per_code_counts: dict[str, int] = {}
                        safe_sheet = _sanitize_filename(sn)
                        for img in ws._images:
                            anchor_row = _get_image_anchor_row(img)
                            if anchor_row is None:
                                continue
                            image_rows_total += 1

                            idx = bisect_right(code_rows, anchor_row) - 1
                            if idx < 0:
                                continue
                            code_row = code_rows[idx]
                            excel_code = row_to_code.get(code_row, "")
                            if not excel_code:
                                continue
                            if first_code_occurrence.get(excel_code) != (sn, code_row):
                                continue

                            pil = _image_to_pil(img)
                            if pil is None:
                                continue

                            # dedupe by image content (hash of PNG bytes)
                            buf = io.BytesIO()
                            pil.convert("RGBA").save(buf, format="PNG")
                            data = buf.getvalue()
                            img_hash = hashlib.sha256(data).hexdigest()

                            code_hashes = per_code_hashes.setdefault(excel_code, set())
                            if img_hash in code_hashes:
                                continue
                            code_hashes.add(img_hash)

                            count = per_code_counts.get(excel_code, 0) + 1
                            per_code_counts[excel_code] = count

                            safe_code = _sanitize_filename(excel_code)
                            out_path = out_dir / f"{safe_sheet}_{safe_code}_{img_hash[:10]}.png"
                            path_str = hash_to_path.get(img_hash)
                            if path_str is None:
                                try:
                                    out_path.write_bytes(data)
                                except Exception:
                                    continue
                                path_str = str(out_path)
                                hash_to_path[img_hash] = path_str

                            lst = image_map.setdefault(excel_code, [])
                            if path_str not in lst:
                                lst.append(path_str)
            except Exception:
                image_map = {}

            total = len(mapping)
            updated = 0
            missing = 0
            missing_codes: list[str] = []
            images_updated = 0
            images_missing = 0
            skipped_existing = 0
            created_new = 0
            i = 0

            # 4) update DB (description + images) using one connection
            conn = self.state.db.connect()
            try:
                for excel_code, desc_pair in mapping.items():
                    i += 1
                    excel_code_str = str(excel_code).strip()

                    # exact match first
                    if excel_code_str in db_code_set:
                        code_to_update = excel_code_str
                    else:
                        # normalized match (only if unique)
                        code_to_update = db_index.get(_normalize_code_soft(excel_code_str), "")

                    if code_to_update:
                        if not should_update_existing(code_to_update):
                            skipped_existing += 1
                            continue
                        desc_vi, desc_en = desc_pair
                        cur = conn.execute(
                            "UPDATE items SET description_excel=?, description_vietnames_from_excel=? WHERE code=?",
                            (str(desc_en).strip(), str(desc_vi).strip(), code_to_update),
                        )
                        if cur.rowcount > 0:
                            updated += 1
                        else:
                            missing += 1
                        missing_codes.append(excel_code_str)
                    else:
                        # Not found in DB: create a new item from Excel fields.
                        desc_vi, desc_en = desc_pair
                        conn.execute(
                            """
                            INSERT INTO items(
                                code,
                                description,
                                description_excel,
                                description_vietnames_from_excel,
                                pdf_path,
                                page,
                                validated,
                                validated_at,
                                category,
                                author,
                                dimension,
                                small_description,
                                shape,
                                blade_tip,
                                surface_treatment,
                                material
                            )
                            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                            """,
                            (
                                excel_code_str,
                                "",
                                str(desc_en).strip(),
                                str(desc_vi).strip(),
                                "",
                                None,
                                0,
                                "",
                                "",
                                "",
                                "",
                                "",
                                "",
                                "",
                                "",
                                "",
                            ),
                        )
                        db_code_set.add(excel_code_str)
                        decision_by_code[excel_code_str] = True
                        created_new += 1

                    # progress update (every 25 rows)
                    if i % 25 == 0:
                        _safe_ui(self.root, lambda i=i, total=total, updated=updated, missing=missing:
                                self._set_status(f"⏳ Cập nhật Excel {i}/{total} | đã cập nhật={updated} | thiếu={missing}"))

                # images: link excel images into assets + item_asset_links (preferred)
                excel_asset_pdf_path = f"excel:{xlsx_path}"
                excel_asset_pdf_path_db = (
                    self.state.db.to_db_path(excel_asset_pdf_path)
                    if self.state.db
                    else excel_asset_pdf_path
                )
                for excel_code, img_paths in image_map.items():
                    # keep order but drop duplicates
                    seen: set[str] = set()
                    unique_paths: list[str] = []
                    for p in img_paths:
                        if p in seen:
                            continue
                        seen.add(p)
                        unique_paths.append(p)
                    excel_code_str = str(excel_code).strip()
                    if excel_code_str in db_code_set:
                        code_to_update = excel_code_str
                    else:
                        code_to_update = db_index.get(_normalize_code_soft(excel_code_str), "")

                    if not code_to_update:
                        # No matching item yet: create a minimal new item for this code.
                        conn.execute(
                            """
                            INSERT INTO items(
                                code,
                                description,
                                description_excel,
                                description_vietnames_from_excel,
                                pdf_path,
                                page,
                                validated,
                                validated_at,
                                category,
                                author,
                                dimension,
                                small_description,
                                shape,
                                blade_tip,
                                surface_treatment,
                                material
                            )
                            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                            """,
                            (
                                excel_code_str,
                                "",
                                "",
                                "",
                                "",
                                None,
                                0,
                                "",
                                "",
                                "",
                                "",
                                "",
                                "",
                                "",
                                "",
                                "",
                            ),
                        )
                        db_code_set.add(excel_code_str)
                        created_new += 1
                        code_to_update = excel_code_str
                        decision_by_code[code_to_update] = True
                    if not should_update_existing(code_to_update):
                        continue

                    row = conn.execute("SELECT id FROM items WHERE code=?", (code_to_update,)).fetchone()
                    if row is None:
                        images_missing += 1
                        continue

                    item_id = int(row["id"])
                    # replace existing asset links so Excel images show in UI
                    conn.execute("DELETE FROM item_asset_links WHERE item_id=?", (item_id,))
                    for idx, p in enumerate(unique_paths):
                        asset_path_db = self.state.db.to_db_path(p) if self.state.db else p
                        asset_row = conn.execute(
                            "SELECT id FROM assets WHERE pdf_path=? AND page=? AND asset_path=?",
                            (excel_asset_pdf_path_db, 0, asset_path_db),
                        ).fetchone()
                        if asset_row:
                            asset_id = int(asset_row["id"])
                        else:
                            cur = conn.execute(
                                """
                                INSERT INTO assets(pdf_path, page, asset_path, x0, y0, x1, y1, source, sha256)
                                VALUES(?,?,?,?,?,?,?,?,?)
                                """,
                                (excel_asset_pdf_path_db, 0, asset_path_db, None, None, None, None, "excel", ""),
                            )
                            asset_id = int(cur.lastrowid)
                        conn.execute(
                            """
                            INSERT OR IGNORE INTO item_asset_links(item_id, asset_id, match_method, score, verified, is_primary)
                            VALUES(?,?,?,?,?,?)
                            """,
                            (item_id, asset_id, "excel", None, 1, 1 if idx == 0 else 0),
                        )
                    images_updated += 1

                conn.commit()
            finally:
                conn.close()

            # 5) refresh UI and show summary
            _safe_ui(self.root, self.refresh_items)
            _safe_ui(self.root, lambda: self._set_status(
                f"✅ Nhập Excel xong | tạo mới={created_new} | đã cập nhật={updated} | bỏ qua={skipped_existing} | thiếu={missing} | ảnh={images_updated}"
            ))
            _safe_ui(
                self.root,
                lambda: messagebox.showinfo(
                    "Nhập Excel xong",
                    "Số dòng đọc: {total}\nTạo mới: {created_new}\nĐã cập nhật: {updated}\nBỏ qua mã đã tồn tại: {skipped_existing}\nMã thiếu: {missing}\n"
                    "Ảnh đã gắn: {images_updated}\nẢnh thiếu: {images_missing}\n"
                    "Ảnh tìm thấy trong file: {image_rows_total}".format(
                        total=total,
                        created_new=created_new,
                        updated=updated,
                        skipped_existing=skipped_existing,
                        missing=missing,
                        images_updated=images_updated,
                        images_missing=images_missing,
                        image_rows_total=image_rows_total,
                    ),
                ),
            )
            if missing_codes:
                _safe_ui(self.root, lambda: self._show_missing_codes(missing_codes))

        self._run_bg("⏳ Đang cập nhật mô tả và ảnh từ Excel...", work)

    def on_search_images_from_excel(self) -> None:
        """
        Load input from the first sheet, match codes to DB items, and write images + descriptions
        into the workbook, plus a formatted Post Processing sheet.
        """
        if not self.state.db:
            messagebox.showwarning("Thiếu CSDL", "Vui lòng tạo/tải CSDL trước (từ PDF).")
            return
        xlsx_path = filedialog.askopenfilename(
            title="Chọn file Excel",
            filetypes=[("Tệp Excel", "*.xlsx *.xls"), ("Tất cả tệp", "*.*")],
        )
        if not xlsx_path:
            return
        xlsx_path = str(xlsx_path)
        ok = self._open_search_options_popup()
        if ok is False:
            return

        def work():
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import PatternFill
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from PIL import Image
            include_desc_en = bool(self.var_export_desc_en.get())
            include_desc_vi = bool(self.var_export_desc_vi.get())
            preserve_image_order = bool(self.var_export_images_in_ui_order.get())

            # Seed only empty database descriptions from the bundled dictionary.
            # Existing user-entered values are never overwritten.
            import_dictionary_into_db(
                self.state.db,
                self.state.project_dir,
            )

            # 1) detect header + code column using existing heuristics on FIRST sheet (input)
            _df, header_row, code_col = detect_excel_code_column(xlsx_path)

            # 2) build DB code indexes (same logic as on_build_excel_db)
            conn = self.state.db.connect()
            try:
                rows = conn.execute("SELECT code FROM items").fetchall()
                db_codes = [str(r["code"]) for r in rows]
            finally:
                conn.close()

            db_code_set = set(db_codes)
            db_index = _build_db_code_index(db_codes)  # normalized -> original db code (unique only)

            items = self.state.db.list_items()
            code_to_item: dict[str, CatalogItem] = {str(it.code): it for it in items}

            # 3) Keep the input sheet and use a temporary worksheet for the
            # legacy layout pass. Only the formatted output sheet is retained.
            wb = load_workbook(xlsx_path)
            input_ws = wb.worksheets[0]
            old_output_names = {"form đầu ra", "post processing"}
            for existing_ws in list(wb.worksheets[1:]):
                if existing_ws.title.strip().casefold() in old_output_names:
                    wb.remove(existing_ws)
            output_ws = wb.create_sheet("_SmartCatalog_Temp", 1)
            header_row_1 = header_row + 1  # openpyxl is 1-based

            # find code column index in INPUT header row
            code_col_idx = None
            for cell in input_ws[header_row_1]:
                if _normalize_header_text(str(cell.value or "")) == _normalize_header_text(code_col):
                    code_col_idx = cell.column
                    break
            if code_col_idx is None:
                raise ValueError(f"Không tìm thấy cột mã '{code_col}' trong dòng tiêu đề Excel.")

            # find qty column (best-effort)
            qty_col_idx = None
            qty_headers = {"qty", "quantity", "so luong", "số lượng", "sl"}
            for cell in input_ws[header_row_1]:
                h = _normalize_header_text(str(cell.value or ""))
                if h in qty_headers or h.startswith("qty"):
                    qty_col_idx = cell.column
                    break

            # read input rows
            input_rows: list[tuple[str, str]] = []
            for r in range(header_row_1 + 1, input_ws.max_row + 1):
                raw_code = input_ws.cell(row=r, column=code_col_idx).value
                excel_code_str = str(raw_code or "").strip()
                if not excel_code_str:
                    continue
                qty_val = ""
                if qty_col_idx is not None:
                    qty_cell = input_ws.cell(row=r, column=qty_col_idx).value
                    qty_val = "" if qty_cell is None else str(qty_cell).strip()
                input_rows.append((excel_code_str, qty_val))

            # Validate before mutating the workbook. A product can have several
            # simultaneous issues (for example, missing EN text and images).
            preflight_items: list[ExportPreflightItem] = []
            for excel_code_str, _qty_val in input_rows:
                if excel_code_str in db_code_set:
                    code_to_match = excel_code_str
                else:
                    code_to_match = db_index.get(_normalize_code_soft(excel_code_str), "")
                item = code_to_item.get(code_to_match) if code_to_match else None
                if item is None:
                    preflight_items.append(
                        ExportPreflightItem(
                            code=excel_code_str,
                            unknown_code=True,
                        )
                    )
                    continue

                description_vi = str(
                    getattr(item, "description_vietnames_from_excel", "") or ""
                ).strip()
                description_en = str(
                    getattr(item, "description_excel", "") or ""
                ).strip()
                pdf_description = str(getattr(item, "description", "") or "").strip()
                effective_en = description_en or pdf_description
                valid_images = [
                    str(path)
                    for path in list(getattr(item, "images", []) or [])
                    if Path(path).is_file()
                ]
                preflight_items.append(
                    ExportPreflightItem(
                        code=str(item.code),
                        item_id=int(item.id),
                        description_vi=description_vi,
                        description_en=description_en,
                        pdf_description=pdf_description,
                        image_paths=valid_images,
                        missing_vi=bool(include_desc_vi and not description_vi),
                        missing_en=bool(include_desc_en and not effective_en),
                        missing_images=not bool(valid_images),
                    )
                )

            issues = [item for item in preflight_items if item.has_issue]
            if issues:
                review_done = threading.Event()
                review_result = {"action": "cancel"}

                def show_preflight_review() -> None:
                    try:
                        review_result["action"] = ExportReviewDialog(
                            self.root,
                            self.state.db,
                            issues,
                        ).show()
                    finally:
                        review_done.set()

                _safe_ui(self.root, show_preflight_review)
                review_done.wait()
                if review_result["action"] != "continue":
                    _safe_ui(self.root, lambda: self._set_status("Đã hủy xuất catalog."))
                    return

                # Descriptions may have changed in the review dialog.
                items = self.state.db.list_items()
                code_to_item = {str(it.code): it for it in items}

            # prepare output sheet: remove merges/images first, then clear rows
            for rng in list(output_ws.merged_cells.ranges):
                if rng.min_row >= 2:
                    try:
                        output_ws.unmerge_cells(str(rng))
                    except KeyError:
                        pass
            if output_ws.max_row > 1:
                output_ws.delete_rows(2, output_ws.max_row - 1)
            if getattr(output_ws, "_images", None):
                output_ws._images = []

            # ensure header exists
            if not output_ws.cell(row=1, column=1).value:
                output_ws.cell(row=1, column=1, value="No.")
                output_ws.cell(row=1, column=2, value="Product Code")
                output_ws.cell(row=1, column=3, value="Product Description")
                output_ws.cell(row=1, column=4, value="Qty.")

            # write output rows + insert images
            updated = 0
            total = 0
            matched = 0
            missing = 0
            missing_codes: list[str] = []
            missing_vi = 0
            missing_en = 0
            missing_images = 0
            out_row = 2
            post_processing_rows: list[PostProcessingRow] = []

            px_to_emu = 9525
            pad = 6

            def col_width_px(col_idx: int) -> int:
                letter = get_column_letter(col_idx)
                w = output_ws.column_dimensions[letter].width
                if w is None:
                    w = output_ws.column_dimensions["A"].width or 8.43
                return int(w * 7 + 5)

            for idx, (excel_code_str, qty_val) in enumerate(input_rows, start=1):
                total += 1
                if excel_code_str in db_code_set:
                    code_to_match = excel_code_str
                else:
                    code_to_match = db_index.get(_normalize_code_soft(excel_code_str), "")

                it = code_to_item.get(code_to_match) if code_to_match else None
                if it:
                    matched += 1
                else:
                    missing += 1
                    missing_codes.append(excel_code_str)

                desc_en_text = ""
                desc_vi_text = ""
                if it:
                    if include_desc_en and getattr(it, "description_excel", ""):
                        desc_en_text = str(it.description_excel or "").strip()
                    if include_desc_vi and getattr(it, "description_vietnames_from_excel", ""):
                        desc_vi_text = str(it.description_vietnames_from_excel or "").strip()
                    if include_desc_en and not desc_en_text and getattr(it, "description", ""):
                        desc_en_text = str(it.description or "").strip()
                    if include_desc_vi and not desc_vi_text:
                        missing_vi += 1
                    if include_desc_en and not desc_en_text:
                        missing_en += 1

                # Decide row ordering: if both EN+VI, show VI above and EN below.
                row1_desc = ""
                row2_desc = ""
                if include_desc_en and include_desc_vi:
                    row1_desc = desc_vi_text
                    row2_desc = desc_en_text
                elif include_desc_vi:
                    row1_desc = desc_vi_text
                elif include_desc_en:
                    row1_desc = desc_en_text

                # data row
                output_ws.cell(row=out_row, column=1, value=idx)
                output_ws.cell(row=out_row, column=2, value=excel_code_str)
                output_ws.cell(row=out_row, column=3, value=row1_desc)
                output_ws.cell(row=out_row, column=4, value=qty_val)
                if row1_desc:
                    line_count = max(1, str(row1_desc).count("\n") + 1)
                    output_ws.row_dimensions[out_row].height = max(15, min(80, 15 * line_count))

                extra_desc_rows = 0
                if include_desc_en and include_desc_vi:
                    extra_desc_rows = 1
                    vn_row = out_row + 1
                    output_ws.cell(row=vn_row, column=3, value=row2_desc)
                    # Keep extra description row compact but readable
                    line_count = max(1, str(row2_desc).count("\n") + 1)
                    output_ws.row_dimensions[vn_row].height = max(15, min(60, 15 * line_count))

                yellow_fill = PatternFill("solid", fgColor="FFFF00")
                primary_missing = bool(
                    (include_desc_vi and not desc_vi_text)
                    if include_desc_vi
                    else (include_desc_en and not desc_en_text)
                )
                secondary_missing = bool(
                    include_desc_vi and include_desc_en and not desc_en_text
                )
                if not it or primary_missing:
                    output_ws.cell(row=out_row, column=3).fill = yellow_fill
                if extra_desc_rows and secondary_missing:
                    output_ws.cell(row=out_row + 1, column=3).fill = yellow_fill

                # image row
                img_row = out_row + 1 + extra_desc_rows
                output_ws.merge_cells(start_row=img_row, start_column=1, end_row=img_row, end_column=4)

                imgs = list(it.images or []) if it else []
                valid_imgs = [str(path) for path in imgs if Path(path).is_file()]
                if it and not valid_imgs:
                    missing_images += 1
                post_processing_rows.append(
                    PostProcessingRow(
                        number=idx,
                        code=excel_code_str,
                        qty=qty_val,
                        desc_primary=row1_desc,
                        desc_secondary=row2_desc,
                        image_paths=list(valid_imgs),
                        highlight_missing_desc=bool(not it or primary_missing),
                        force_secondary_row=bool(include_desc_vi and include_desc_en),
                        highlight_missing_secondary=secondary_missing,
                    )
                )
                if valid_imgs:
                    updated += 1

                    # Load images and compute base sizes
                    loaded: list[tuple[Image.Image, int, int]] = []
                    for img_path in valid_imgs:
                        try:
                            p = Path(img_path)
                            if not p.exists():
                                continue
                            pil = Image.open(p).convert("RGBA")
                            if not preserve_image_order and pil.height > pil.width:
                                pil = pil.rotate(90, expand=True)
                            w, h = pil.size
                            if h <= 0 or w <= 0:
                                continue
                            loaded.append((pil, w, h))
                        except Exception:
                            continue

                    if loaded:
                        if not preserve_image_order:
                            loaded.sort(key=lambda t: t[1] * t[2])

                        total_w = sum(w for _, w, _ in loaded) + pad * max(0, len(loaded) - 1)
                        max_h = max(h for _, _, h in loaded)

                        min_h = max_h + 20
                        output_ws.row_dimensions[img_row].height = max(min_h, max_h + 6) / 1.333

                        available_w = sum(col_width_px(c) for c in range(1, 5))
                        if available_w < 1:
                            available_w = total_w
                        x_off = max(0, int((available_w - total_w) / 2))
                        col_widths = [col_width_px(c) for c in range(1, 5)]
                        v_pad = 10
                        row_height_px = int((output_ws.row_dimensions[img_row].height or 0) * 1.333)
                        target_h = max(1, row_height_px - (v_pad * 2))

                        scale_h = 1.0 if max_h <= target_h else (target_h / float(max_h))
                        scale_w = 1.0 if total_w <= available_w else (available_w / float(total_w))
                        scale = min(1.0, scale_h, scale_w)

                        for pil, w, h in loaded:
                            try:
                                new_w = max(1, int(w * scale))
                                new_h = max(1, int(h * scale))

                                buf = io.BytesIO()
                                if new_w != w or new_h != h:
                                    pil_resized = pil.resize((new_w, new_h), Image.LANCZOS)
                                else:
                                    pil_resized = pil
                                pil_resized.save(buf, format="PNG")
                                buf.seek(0)
                                xl_img = XLImage(buf)
                                xl_img.width = new_w
                                xl_img.height = new_h

                                col_idx = 0
                                col_off = x_off
                                while col_idx < len(col_widths) - 1 and col_off >= col_widths[col_idx]:
                                    col_off -= col_widths[col_idx]
                                    col_idx += 1

                                row_height_px = int((output_ws.row_dimensions[img_row].height or 0) * 1.333)
                                y_off = 0
                                if row_height_px > new_h + v_pad * 2:
                                    y_off = int((row_height_px - new_h) / 2)
                                elif row_height_px > new_h:
                                    y_off = v_pad

                                marker = AnchorMarker(
                                    col=col_idx,
                                    colOff=int(col_off * px_to_emu),
                                    row=img_row - 1,
                                    rowOff=int(y_off * px_to_emu),
                                )
                                ext = XDRPositiveSize2D(new_w * px_to_emu, new_h * px_to_emu)
                                xl_img.anchor = OneCellAnchor(_from=marker, ext=ext)
                                output_ws.add_image(xl_img)

                                x_off += new_w + pad
                            except Exception:
                                continue

                out_row += 2 + extra_desc_rows

            write_post_processing_sheet(
                wb,
                post_processing_rows,
                preserve_image_order=preserve_image_order,
                sheet_name=POST_PROCESSING_SHEET_NAME,
                index=wb.worksheets.index(output_ws) + 1,
            )
            wb.remove(output_ws)
            wb.save(xlsx_path)
            apply_post_processing_branding(
                xlsx_path,
                project_dir=self.state.project_dir,
            )

            _safe_ui(
                self.root,
                lambda: self._show_export_result(
                    xlsx_path=xlsx_path,
                    matched=matched,
                    total=total,
                    images_found=updated,
                    missing_vi=missing_vi,
                    missing_en=missing_en,
                    missing_images=missing_images,
                    missing_codes=missing_codes,
                ),
            )
            _safe_ui(
                self.root,
                lambda: self._set_status(
                    f"✅ Xuất Excel: khớp {matched}/{total} | thiếu VI={missing_vi}, "
                    f"EN={missing_en}, ảnh={missing_images}, mã={missing} | "
                    f"tạo {POST_PROCESSING_SHEET_NAME}"
                )
            )

        self._run_bg("⏳ Extracting images by code...", work)


###################################################################################################
def create_main_window(root: tk.Tk, state: Optional[AppState] = None) -> MainWindow:
    return MainWindow(root, state=state)
