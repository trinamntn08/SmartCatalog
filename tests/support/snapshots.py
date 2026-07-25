from __future__ import annotations

import hashlib
import json
import sqlite3
import zipfile
from contextlib import closing
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(64 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def snapshot_sqlite(path: Path) -> dict[str, Any]:
    """Return a stable schema-and-data representation of a SQLite database."""

    with closing(sqlite3.connect(path)) as conn:
        conn.row_factory = sqlite3.Row
        table_rows = conn.execute(
            """
            SELECT name, sql
            FROM sqlite_master
            WHERE type='table' AND name NOT LIKE 'sqlite_%'
            ORDER BY name
            """
        ).fetchall()

        tables: dict[str, Any] = {}
        for table_row in table_rows:
            table_name = str(table_row["name"])
            quoted_name = '"' + table_name.replace('"', '""') + '"'
            columns = [
                dict(row)
                for row in conn.execute(f"PRAGMA table_info({quoted_name})").fetchall()
            ]
            data = [
                list(row)
                for row in conn.execute(
                    f"SELECT * FROM {quoted_name} ORDER BY rowid"
                ).fetchall()
            ]
            tables[table_name] = {
                "sql": str(table_row["sql"] or ""),
                "columns": columns,
                "rows": data,
            }

    return {"tables": tables}


def snapshot_workbook(path: Path) -> dict[str, Any]:
    """Capture stable workbook semantics while ignoring ZIP timestamps."""

    workbook = load_workbook(path, data_only=False)
    try:
        sheets: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            cells = []
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cells.append([cell.coordinate, cell.value])
            sheets.append(
                {
                    "title": sheet.title,
                    "cells": cells,
                    "merged_ranges": sorted(str(rng) for rng in sheet.merged_cells.ranges),
                    "image_count": len(getattr(sheet, "_images", [])),
                }
            )
        return {"sheets": sheets}
    finally:
        workbook.close()


def snapshot_xlsx_package(path: Path) -> dict[str, str]:
    """Hash XLSX package members by content, independent of ZIP metadata."""

    with zipfile.ZipFile(path, "r") as archive:
        return {
            name: hashlib.sha256(archive.read(name)).hexdigest()
            for name in sorted(archive.namelist())
        }


def write_snapshot(path: Path, value: dict[str, Any]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    return path
