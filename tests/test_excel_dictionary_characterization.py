from __future__ import annotations

import csv
import hashlib
import io
import unittest
from contextlib import closing
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

from tests import support as _test_support  # Ensures src/ is importable.
from tests.support.fixtures import TemporaryProject, create_image_fixture

from smartcatalog.db.catalog_db import CatalogDB
from smartcatalog.loader.excel_loader import (
    detect_excel_code_column,
    load_code_to_description_from_excel,
    load_code_to_vi_en_from_excel,
    normalize_code_soft,
)
from smartcatalog.services.excel_catalog_import import import_excel_catalog
from smartcatalog.ui.main_window import (
    MainWindow,
    _build_db_code_index,
    _get_image_anchor_row,
    _image_to_pil,
    _sanitize_filename,
)
from smartcatalog.utils.description_dictionary import (
    import_dictionary_into_db,
    read_description_dictionary,
    resolve_description_dictionary,
)


def save_rows(path: Path, sheets: dict[str, list[list[object]]]) -> Path:
    workbook = Workbook()
    first = True
    for title, rows in sheets.items():
        sheet = workbook.active if first else workbook.create_sheet()
        first = False
        sheet.title = title
        for row in rows:
            sheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


class ExcelLoaderCharacterizationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.project = TemporaryProject()
        self.project.__enter__()
        self.addCleanup(self.project.__exit__, None, None, None)

    def test_detects_header_after_title_rows_and_bilingual_following_row(self) -> None:
        path = save_rows(
            self.project.path("delayed-header.xlsx"),
            {
                "Catalog": [
                    ["Medical catalog"],
                    ["Generated for review"],
                    ["No.", "Product Code", "Product Description", "Qty."],
                    [1, "12-345-67", "Mô tả tiếng Việt", 2],
                    [None, None, "English description", None],
                    [2, "98-765-43", "Chỉ tiếng Việt", 1],
                ]
            },
        )

        dataframe, header_row, code_column = detect_excel_code_column(path)
        bilingual = load_code_to_vi_en_from_excel(path)

        self.assertEqual(header_row, 2)
        self.assertEqual(code_column, "Product Code")
        self.assertEqual(list(dataframe.columns), ["No.", "Product Code", "Product Description", "Qty."])
        self.assertEqual(
            bilingual,
            {
                "12-345-67": ("Mô tả tiếng Việt", "English description"),
                "98-765-43": ("Chỉ tiếng Việt", ""),
            },
        )

    def test_accepts_alternate_ascii_headers_and_skips_junk_rows(self) -> None:
        path = save_rows(
            self.project.path("alternate-headers.xlsx"),
            {
                "Catalog": [
                    ["Report"],
                    ["Item Code", "Name"],
                    ["not a product", "Title row"],
                    ["12 345 67", "First"],
                    ["98-765-43", ""],
                    [None, "Missing code"],
                    ["98–765–43", "Unicode dash"],
                ]
            },
        )

        mapping = load_code_to_description_from_excel(path)

        self.assertEqual(
            mapping,
            {
                "12 345 67": "First",
                "98–765–43": "Unicode dash",
            },
        )

    def test_sheet_selection_is_explicit_and_first_sheet_is_default(self) -> None:
        path = save_rows(
            self.project.path("multiple-sheets.xlsx"),
            {
                "First": [
                    ["Product Code", "Description"],
                    ["12-345-67", "First sheet"],
                ],
                "Second": [
                    ["Product Code", "Description"],
                    ["12-345-67", "Second sheet"],
                    ["98-765-43", "Second product"],
                ],
                "Empty": [],
                "Unsupported": [["SKU", "Details"], ["12-345-67", "Ignored"]],
            },
        )

        self.assertEqual(
            load_code_to_description_from_excel(path),
            {"12-345-67": "First sheet"},
        )
        self.assertEqual(
            load_code_to_description_from_excel(path, sheet_name="Second"),
            {
                "12-345-67": "Second sheet",
                "98-765-43": "Second product",
            },
        )
        with self.assertRaises(ValueError):
            load_code_to_description_from_excel(path, sheet_name="Empty")
        with self.assertRaises(ValueError):
            load_code_to_description_from_excel(path, sheet_name="Unsupported")

    def test_duplicate_codes_use_the_last_row_within_one_sheet(self) -> None:
        path = save_rows(
            self.project.path("duplicates.xlsx"),
            {
                "Catalog": [
                    ["Product Code", "Description"],
                    ["12-345-67", "First value"],
                    ["12-345-67", "Last value"],
                ]
            },
        )

        self.assertEqual(
            load_code_to_description_from_excel(path),
            {"12-345-67": "Last value"},
        )

    def test_normalized_index_keeps_only_unique_mappings(self) -> None:
        self.assertEqual(normalize_code_soft(" 12 – 345 — 67 "), "12-345-67")
        index = _build_db_code_index(
            ["12-345-67", "98-765-43", "98–765–43", "77 777 77"]
        )

        self.assertEqual(index["12-345-67"], "12-345-67")
        self.assertNotIn("98-765-43", index)
        self.assertEqual(index["7777777"], "77 777 77")


class EmbeddedWorkbookImageCharacterizationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.project = TemporaryProject()
        self.project.__enter__()
        self.addCleanup(self.project.__exit__, None, None, None)

    def create_workbook_with_images(self) -> tuple[Path, Path, Path]:
        red = create_image_fixture(
            self.project.path("images", "red.png"),
            size=(18, 12),
            color=(200, 20, 20),
        )
        blue = create_image_fixture(
            self.project.path("images", "blue.png"),
            size=(12, 18),
            color=(20, 20, 200),
        )
        workbook_path = self.project.path("embedded-images.xlsx")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Catalog Sheet"
        sheet.append(["Product Code", "Product Description"])
        sheet.append(["12-345-67", "VI"])
        sheet.append([None, "EN"])
        sheet.append(["98-765-43", "VI 2"])

        image_streams: list[io.BytesIO] = []
        for source, anchor in (
            (red, "C1"),
            (red, "C2"),
            (red, "D3"),
            (blue, "C4"),
        ):
            stream = io.BytesIO(source.read_bytes())
            image_streams.append(stream)
            sheet.add_image(XLImage(stream), anchor)

        workbook.save(workbook_path)
        workbook.close()
        for stream in image_streams:
            stream.close()
        return workbook_path, red, blue

    def test_embedded_images_reopen_with_stable_anchors_and_content(self) -> None:
        workbook_path, _red, _blue = self.create_workbook_with_images()
        workbook = load_workbook(workbook_path)
        try:
            images = list(workbook["Catalog Sheet"]._images)
            anchors = [_get_image_anchor_row(image) for image in images]
            hashes = []
            for embedded in images:
                pil_image = _image_to_pil(embedded)
                self.assertIsNotNone(pil_image)
                try:
                    buffer = io.BytesIO()
                    pil_image.convert("RGBA").save(buffer, format="PNG")
                    hashes.append(hashlib.sha256(buffer.getvalue()).hexdigest())
                finally:
                    pil_image.close()
        finally:
            workbook.close()

        self.assertEqual(anchors, [1, 2, 3, 4])
        self.assertEqual(len(set(hashes[:3])), 1)
        self.assertNotEqual(hashes[0], hashes[3])

    def test_current_filename_sanitization_is_deterministic(self) -> None:
        self.assertEqual(_sanitize_filename("Catalog Sheet"), "Catalog_Sheet")
        self.assertEqual(_sanitize_filename("12 / 345 / 67"), "12_345_67")
        self.assertEqual(_sanitize_filename("***"), "item")


class ExcelImportWorkflowCharacterizationTests(unittest.TestCase):
    class ImmediateRoot:
        @staticmethod
        def after(_delay: int, callback) -> None:
            callback()

    def setUp(self) -> None:
        self.project = TemporaryProject()
        self.project.__enter__()
        self.addCleanup(self.project.__exit__, None, None, None)
        self.data_dir = self.project.create_runtime_layout()
        self.db = CatalogDB(
            self.data_dir / "sql" / "catalog.db",
            data_dir=self.data_dir,
        )

    def create_import_workbook(self) -> Path:
        repeated = create_image_fixture(
            self.project.path("source-images", "repeated.png"),
            color=(180, 40, 40),
        )
        unique = create_image_fixture(
            self.project.path("source-images", "unique.png"),
            color=(40, 40, 180),
        )
        path = self.project.path("catalog-import.xlsx")
        workbook = Workbook()
        first = workbook.active
        first.title = "First Sheet"
        first.append(["Title"])
        first.append(["Product Code", "Product Description"])
        first.append(["12-345-67", "VI exact"])
        first.append([None, "EN exact"])
        first.append(["98 – 765 – 43", "VI normalized"])
        first.append([None, "EN normalized"])
        first.append(["77-777-77", "VI new"])
        first.append([None, "EN new"])
        image_streams: list[io.BytesIO] = []

        def add_image(sheet, source: Path, anchor: str) -> None:
            stream = io.BytesIO(source.read_bytes())
            image_streams.append(stream)
            sheet.add_image(XLImage(stream), anchor)

        add_image(first, repeated, "C3")
        add_image(first, repeated, "D4")
        add_image(first, unique, "C5")

        second = workbook.create_sheet("Second Sheet")
        second.append(["Product Code", "Product Description"])
        second.append(["12-345-67", "VI duplicate"])
        second.append([None, "EN duplicate"])
        add_image(second, unique, "C2")

        workbook.create_sheet("Empty")
        unsupported = workbook.create_sheet("Unsupported")
        unsupported.append(["SKU", "Details"])
        unsupported.append(["66-666-66", "Ignored"])

        workbook.save(path)
        workbook.close()
        for stream in image_streams:
            stream.close()
        return path

    def test_headless_import_preserves_current_matching_image_and_sheet_rules(self) -> None:
        self.db.upsert_by_code(code="12-345-67", page=1)
        self.db.upsert_by_code(code="98-765-43", page=2)
        workbook_path = self.create_import_workbook()

        window = SimpleNamespace(
            state=SimpleNamespace(
                db=self.db,
                assets_dir=self.data_dir / "assets",
            ),
            root=self.ImmediateRoot(),
            _run_bg=lambda _title, work: work(),
            _ask_update_existing_item_dialog=lambda _code: (True, True),
            refresh_items=lambda: None,
            _set_status=lambda _message: None,
            _show_missing_codes=lambda _codes: None,
        )

        with (
            patch(
                "smartcatalog.ui.main_window.filedialog.askopenfilename",
                return_value=str(workbook_path),
            ),
            patch("smartcatalog.ui.main_window.messagebox.showinfo"),
            patch("smartcatalog.ui.main_window.messagebox.showwarning"),
            patch("smartcatalog.ui.main_window.messagebox.showerror"),
        ):
            MainWindow.on_build_excel_db(window)

        exact = self.db.get_item_by_code("12-345-67")
        normalized = self.db.get_item_by_code("98-765-43")
        created = self.db.get_item_by_code("77-777-77")

        self.assertEqual(
            (
                exact.description_vietnames_from_excel,
                exact.description_excel,
            ),
            ("VI exact", "EN exact"),
        )
        self.assertEqual(
            (
                normalized.description_vietnames_from_excel,
                normalized.description_excel,
            ),
            ("VI normalized", "EN normalized"),
        )
        self.assertIsNotNone(created)
        self.assertEqual(
            (
                created.description_vietnames_from_excel,
                created.description_excel,
            ),
            ("VI new", "EN new"),
        )

        exact_sources = self.db.list_image_sources_for_item(exact.id)
        normalized_sources = self.db.list_image_sources_for_item(normalized.id)
        self.assertEqual(len(exact_sources), 1)
        self.assertEqual(len(normalized_sources), 1)
        self.assertEqual(exact_sources[0][1], "excel")
        self.assertEqual(normalized_sources[0][1], "excel")
        self.assertTrue(Path(exact_sources[0][0]).is_file())
        self.assertTrue(Path(normalized_sources[0][0]).is_file())
        self.assertIn("First_Sheet_12-345-67_", Path(exact_sources[0][0]).name)
        self.assertIsNone(self.db.get_item_by_code("66-666-66"))

        with closing(self.db.connect()) as conn:
            asset_count = conn.execute("SELECT COUNT(*) FROM assets").fetchone()[0]
            links = conn.execute(
                """
                SELECT a.source, l.match_method, l.verified, l.is_primary
                FROM item_asset_links l
                JOIN assets a ON a.id = l.asset_id
                ORDER BY l.id
                """
            ).fetchall()
        self.assertEqual(asset_count, 2)
        self.assertEqual(
            [
                (
                    row["source"],
                    row["match_method"],
                    row["verified"],
                    row["is_primary"],
                )
                for row in links
            ],
            [
                ("excel", "excel", 1, 1),
                ("excel", "excel", 1, 1),
            ],
        )

    def test_extracted_service_preserves_import_result_and_database_shape(self) -> None:
        self.db.upsert_by_code(code="12-345-67", page=1)
        self.db.upsert_by_code(code="98-765-43", page=2)
        workbook_path = self.create_import_workbook()
        decisions: list[str] = []

        result = import_excel_catalog(
            workbook_path,
            db=self.db,
            assets_dir=self.data_dir / "assets",
            should_update_existing=lambda code: decisions.append(code) or True,
        )

        self.assertEqual(
            (
                result.total,
                result.created_new,
                result.updated,
                result.skipped_existing,
                result.images_updated,
                result.images_missing,
                result.image_rows_total,
            ),
            (3, 1, 2, 0, 2, 0, 4),
        )
        self.assertEqual(result.missing_codes, ("12-345-67", "98 – 765 – 43"))
        self.assertEqual(decisions, ["12-345-67", "98-765-43", "12-345-67", "98-765-43"])

        with closing(self.db.connect()) as connection:
            items = connection.execute(
                """
                SELECT code, description_excel, description_vietnames_from_excel
                FROM items ORDER BY id
                """
            ).fetchall()
            assets = connection.execute(
                "SELECT source, page FROM assets ORDER BY id"
            ).fetchall()
            links = connection.execute(
                """
                SELECT l.match_method, l.verified, l.is_primary
                FROM item_asset_links l ORDER BY l.id
                """
            ).fetchall()
        self.assertEqual(
            [tuple(row) for row in items],
            [
                ("12-345-67", "EN exact", "VI exact"),
                ("98-765-43", "EN normalized", "VI normalized"),
                ("77-777-77", "EN new", "VI new"),
            ],
        )
        self.assertEqual([tuple(row) for row in assets], [("excel", 0), ("excel", 0)])
        self.assertEqual(
            [tuple(row) for row in links],
            [("excel", 1, 1), ("excel", 1, 1)],
        )


class DictionaryCharacterizationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.project = TemporaryProject()
        self.project.__enter__()
        self.addCleanup(self.project.__exit__, None, None, None)
        self.data_dir = self.project.create_runtime_layout()
        self.db = CatalogDB(
            self.data_dir / "sql" / "catalog.db",
            data_dir=self.data_dir,
        )

    def write_dictionary(self, rows: list[list[str]], *, bom: bool = True) -> Path:
        path = self.data_dir / "dictionary.csv"
        encoding = "utf-8-sig" if bom else "utf-8"
        with path.open("w", encoding=encoding, newline="") as stream:
            writer = csv.writer(stream)
            writer.writerows(rows)
        return path

    def test_resolve_returns_only_existing_runtime_dictionary(self) -> None:
        self.assertIsNone(resolve_description_dictionary(self.project.path()))
        path = self.write_dictionary([["12-345-67", "VI", "EN"]])
        self.assertEqual(
            resolve_description_dictionary(self.project.path()),
            path,
        )

    def test_utf8_bom_reader_trims_values_skips_short_rows_and_keeps_last_duplicate(self) -> None:
        path = self.write_dictionary(
            [
                [" code ", " Mô tả ", " Description "],
                ["short", "missing english"],
                ["", "blank code", "ignored"],
                ["code", "Giá trị cuối", "Last value"],
            ]
        )

        self.assertEqual(
            read_description_dictionary(path),
            {"code": ("Giá trị cuối", "Last value")},
        )

    def test_import_fills_empty_fields_without_overwriting_user_values(self) -> None:
        self.db.upsert_by_code(
            code="12-345-67",
            page=1,
            description_excel="",
            description_vietnames_from_excel="User VI",
        )
        self.db.upsert_by_code(
            code="98-765-43",
            page=2,
            description_excel="",
            description_vietnames_from_excel="",
        )
        self.write_dictionary(
            [
                ["12-345-67", "Dictionary VI", "Dictionary EN"],
                ["98-765-43", "Second VI", "Second EN"],
                ["77-777-77", "Missing VI", "Missing EN"],
            ]
        )

        matched = import_dictionary_into_db(self.db, self.project.path())
        first = self.db.get_item_by_code("12-345-67")
        second = self.db.get_item_by_code("98-765-43")

        self.assertEqual(matched, 2)
        self.assertEqual(
            (first.description_vietnames_from_excel, first.description_excel),
            ("User VI", "Dictionary EN"),
        )
        self.assertEqual(
            (second.description_vietnames_from_excel, second.description_excel),
            ("Second VI", "Second EN"),
        )
        self.assertIsNone(self.db.get_item_by_code("77-777-77"))

    def test_import_without_dictionary_returns_zero(self) -> None:
        self.db.upsert_by_code(code="12-345-67", page=1)
        self.assertEqual(import_dictionary_into_db(self.db, self.project.path()), 0)


if __name__ == "__main__":
    unittest.main()
