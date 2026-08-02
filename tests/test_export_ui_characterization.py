from __future__ import annotations

import io
import unittest
import zipfile
from contextlib import closing
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from PIL import Image

from tests import support as _test_support  # Ensures src/ is importable.
from tests.support.fixtures import TemporaryProject, create_image_fixture
from tests.support.snapshots import snapshot_workbook, snapshot_xlsx_package

from smartcatalog.db.catalog_db import CatalogDB
from smartcatalog.services.catalog_export import (
    CatalogExportOptions,
    export_catalog,
)
from smartcatalog.services.export_preflight import (
    ExportPreflightItem as ServiceExportPreflightItem,
    prepare_export_preflight,
    save_export_preflight_changes,
    save_export_preflight_descriptions,
)
from smartcatalog.services.workbook_product_reader import read_workbook_products
from smartcatalog.ui.export_review_dialog import ExportPreflightItem, ExportReviewDialog
from smartcatalog.ui.main_window import MainWindow
from smartcatalog.utils.post_processing import (
    DEFAULT_SHEET_NAME,
    PostProcessingRow,
    _load_images,
    apply_post_processing_branding,
    write_post_processing_sheet,
)


class BooleanValue:
    def __init__(self, value: bool) -> None:
        self.value = value

    def get(self) -> bool:
        return self.value


class ImmediateRoot:
    @staticmethod
    def after(_delay: int, callback) -> None:
        callback()


class ExportTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.project = TemporaryProject()
        self.project.__enter__()
        self.addCleanup(self.project.__exit__, None, None, None)
        self.data_dir = self.project.create_runtime_layout()
        self.db = CatalogDB(
            self.data_dir / "sql" / "catalog.db",
            data_dir=self.data_dir,
        )

    def create_input_workbook(
        self,
        rows: list[tuple[str, object]],
        *,
        include_old_output: bool = False,
    ) -> Path:
        path = self.project.path("export-input.xlsx")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Input"
        sheet.append(["Requested products"])
        sheet.append(["Product Code", "Qty."])
        for code, quantity in rows:
            sheet.append([code, quantity])
        if include_old_output:
            old = workbook.create_sheet("Post Processing")
            old["A1"] = "old output"
        workbook.save(path)
        workbook.close()
        return path

    def create_item(
        self,
        *,
        code: str,
        description_vi: str,
        description_en: str,
        pdf_description: str = "",
        image_specs: list[tuple[str, tuple[int, int], tuple[int, int, int]]] | None = None,
    ) -> int:
        item_id = self.db.upsert_by_code(
            code=code,
            page=1,
            description=pdf_description,
            description_excel=description_en,
            description_vietnames_from_excel=description_vi,
        )
        for index, (name, size, color) in enumerate(image_specs or []):
            image_path = create_image_fixture(
                self.data_dir / "assets" / "manual_import" / name,
                size=size,
                color=color,
            )
            asset_id = self.db.upsert_asset(
                pdf_path="",
                page=1,
                asset_path=str(image_path),
                source="add",
            )
            self.db.link_asset_to_item(
                item_id=item_id,
                asset_id=asset_id,
                match_method="manual",
                verified=True,
                is_primary=(index == 0),
            )
        return item_id

    def run_export(
        self,
        workbook_path: Path,
        *,
        include_vi: bool,
        include_en: bool,
        preserve_image_order: bool,
        review_action: str = "continue",
    ) -> tuple[list[dict], list[str], object]:
        results: list[dict] = []
        statuses: list[str] = []
        captured_issues: list[list[ExportPreflightItem]] = []

        class ReviewDialogStub:
            def __init__(stub_self, _root, _db, issues):
                captured_issues.append(list(issues))

            def show(stub_self) -> str:
                return review_action

        window = SimpleNamespace(
            state=SimpleNamespace(
                db=self.db,
                project_dir=self.project.path(),
            ),
            root=ImmediateRoot(),
            var_export_desc_en=BooleanValue(include_en),
            var_export_desc_vi=BooleanValue(include_vi),
            var_export_images_in_ui_order=BooleanValue(preserve_image_order),
            _open_search_options_popup=lambda: True,
            _run_bg=lambda _title, work: work(),
            _set_status=lambda message: statuses.append(message),
            _show_export_result=lambda **kwargs: results.append(kwargs),
            refresh_items=lambda: None,
        )

        with (
            patch(
                "smartcatalog.ui.controllers.workflows_controller."
                "filedialog.askopenfilename",
                return_value=str(workbook_path),
            ),
            patch(
                "smartcatalog.ui.controllers.workflows_controller.ExportReviewDialog",
                ReviewDialogStub,
            ),
        ):
            MainWindow.on_export_catalog(window)

        return results, statuses, captured_issues


class PreflightCharacterizationTests(unittest.TestCase):
    def test_ui_reexports_service_preflight_item(self) -> None:
        self.assertIs(ExportPreflightItem, ServiceExportPreflightItem)

    def test_issue_flags_and_status_text_follow_current_priority(self) -> None:
        ready = ExportPreflightItem(code="12-345-67")
        issue = ExportPreflightItem(
            code="98-765-43",
            missing_vi=True,
            missing_en=True,
            missing_images=True,
            unknown_code=True,
        )

        self.assertFalse(ready.has_issue)
        self.assertTrue(issue.has_issue)
        status = issue.status_text()
        self.assertLess(status.index("VI"), status.index("EN"))
        self.assertIn("98", issue.code)

    def test_review_editor_loads_existing_content_and_images(self) -> None:
        class ValueSink:
            value = ""

            def set(self, value: str) -> None:
                self.value = value

        class TextSink:
            value = ""

            def configure(self, **_kwargs) -> None:
                pass

            def delete(self, *_args) -> None:
                self.value = ""

            def insert(self, _index: str, value: str) -> None:
                self.value = value

            def edit_modified(self, _value: bool) -> None:
                pass

        dialog = ExportReviewDialog.__new__(ExportReviewDialog)
        dialog.code_var = ValueSink()
        dialog.status_var = ValueSink()
        dialog.vi_text = TextSink()
        dialog.en_text = TextSink()
        dialog.save_button = SimpleNamespace(configure=lambda **_kwargs: None)
        dialog.add_image_button = SimpleNamespace(configure=lambda **_kwargs: None)
        displayed_images: list[list[str]] = []
        dialog._refresh_image_panel = (
            lambda paths: displayed_images.append(list(paths))
        )

        dialog._load_issue(
            ExportPreflightItem(
                code="12-345-67",
                description_vi="",
                description_en="",
                pdf_description="Existing PDF English",
                image_paths=["first.png", "second.png"],
                missing_vi=True,
            )
        )

        self.assertEqual(dialog.vi_text.value, "")
        self.assertEqual(dialog.en_text.value, "Existing PDF English")
        self.assertEqual(displayed_images, [["first.png", "second.png"]])

    def test_review_add_and_remove_images_remain_drafts(self) -> None:
        issue = ExportPreflightItem(
            code="12-345-67",
            missing_images=True,
        )
        statuses: list[str] = []
        previews: list[list[str]] = []
        dialog = ExportReviewDialog.__new__(ExportReviewDialog)
        dialog._selected = issue
        dialog.window = object()
        dialog.status_var = SimpleNamespace(
            set=lambda value: statuses.append(value)
        )
        dialog._dirty = False
        dialog._refresh_image_panel = (
            lambda paths, **_kwargs: previews.append(list(paths))
        )

        with (
            patch(
                "smartcatalog.ui.export_review_dialog.filedialog.askopenfilenames",
                return_value=("first.png", "second.png"),
            ),
            patch(
                "smartcatalog.ui.export_review_dialog."
                "validate_export_preflight_image",
                side_effect=lambda path: path,
            ),
        ):
            dialog._add_images()

        self.assertEqual(issue.image_paths, ["first.png", "second.png"])
        self.assertEqual(issue.persisted_image_paths, [])
        self.assertTrue(issue.missing_images)
        self.assertTrue(dialog._dirty)
        self.assertEqual(previews, [["first.png", "second.png"]])
        self.assertTrue(statuses)

        dialog._drag_source_index = 1
        dialog._drag_active = True
        dialog.image_thumbs_canvas = SimpleNamespace(
            configure=lambda **_kwargs: None
        )
        dialog._image_index_at = lambda _x, _y: 0
        dialog._on_image_drag_release(SimpleNamespace(x_root=10, y_root=10))

        self.assertEqual(issue.image_paths, ["second.png", "first.png"])
        self.assertEqual(previews[-1], ["second.png", "first.png"])

        dialog._selected_image_index = 0
        dialog._remove_image()

        self.assertEqual(issue.image_paths, ["first.png"])
        self.assertEqual(issue.persisted_image_paths, [])
        self.assertEqual(previews[-1], ["first.png"])


class ExportServiceBoundaryTests(ExportTestCase):
    def test_preflight_save_replaces_multiple_images_only_when_called(self) -> None:
        item_id = self.create_item(
            code="77-777-77",
            description_vi="",
            description_en="",
        )
        path = self.create_input_workbook([("77-777-77", 1)])
        first_source = create_image_fixture(
            self.project.path("first-selected.png"),
            size=(80, 60),
            color=(20, 40, 60),
        )
        second_source = create_image_fixture(
            self.project.path("second-selected.png"),
            size=(60, 80),
            color=(60, 40, 20),
        )
        rows = read_workbook_products(path)
        issue = prepare_export_preflight(
            rows,
            db=self.db,
            include_description_vi=False,
            include_description_en=False,
        ).issues[0]

        issue.image_paths.extend([str(first_source), str(second_source)])
        self.assertEqual(self.db.get_item_by_code("77-777-77").images, [])

        save_export_preflight_changes(
            issue,
            db=self.db,
            description_vi="",
            description_en="",
            image_paths=list(issue.image_paths),
        )

        stored_paths = list(issue.image_paths)
        self.assertEqual(len(stored_paths), 2)
        for stored_path in stored_paths:
            stored = Path(stored_path)
            self.assertTrue(stored.is_file())
            self.assertEqual(
                stored.parent,
                self.data_dir / "assets" / "manual_import",
            )
        self.assertEqual(issue.persisted_image_paths, stored_paths)
        self.assertFalse(issue.missing_images)
        self.assertFalse(issue.has_issue)
        saved = self.db.get_item_by_code("77-777-77")
        self.assertEqual(saved.images, stored_paths)
        self.assertEqual(
            self.db.list_image_sources_for_item(item_id),
            [(stored_paths[0], "add"), (stored_paths[1], "add")],
        )
        links = self.db.list_asset_links_for_item(item_id)
        self.assertEqual(len(links), 2)
        self.assertEqual([link["is_primary"] for link in links], [1, 0])

        reordered_paths = [stored_paths[1], stored_paths[0]]
        save_export_preflight_changes(
            issue,
            db=self.db,
            description_vi="",
            description_en="",
            image_paths=reordered_paths,
        )

        self.assertEqual(issue.image_paths, reordered_paths)
        self.assertEqual(
            self.db.get_item_by_code("77-777-77").images,
            reordered_paths,
        )

        save_export_preflight_changes(
            issue,
            db=self.db,
            description_vi="",
            description_en="",
            image_paths=[reordered_paths[0]],
        )

        self.assertEqual(issue.image_paths, [reordered_paths[0]])
        self.assertEqual(
            self.db.get_item_by_code("77-777-77").images,
            [reordered_paths[0]],
        )
        remaining_link = self.db.list_asset_links_for_item(item_id)[0]
        self.assertEqual(remaining_link["is_primary"], 1)

        result = export_catalog(
            path,
            rows,
            db=self.db,
            project_dir=self.project.path(),
            options=CatalogExportOptions(
                include_description_vi=False,
                include_description_en=False,
                preserve_image_order=True,
            ),
        )
        self.assertEqual(result.images_found, 1)
        self.assertEqual(result.missing_images, 0)

    def test_preflight_save_creates_unknown_item_for_current_export(self) -> None:
        issue = ExportPreflightItem(code="77-777-77", unknown_code=True)

        save_export_preflight_descriptions(
            issue,
            db=self.db,
            description_vi="Mô tả mới",
            description_en="New description",
        )

        saved = self.db.get_item_by_code("77-777-77")
        self.assertIsNotNone(saved)
        self.assertEqual(saved.description_vietnames_from_excel, "Mô tả mới")
        self.assertEqual(saved.description_excel, "New description")
        self.assertFalse(issue.unknown_code)
        self.assertEqual(issue.item_id, saved.id)
        self.assertTrue(issue.missing_images)

        path = self.create_input_workbook([("77-777-77", 1)])
        rows = read_workbook_products(path)
        result = export_catalog(
            path,
            rows,
            db=self.db,
            project_dir=self.project.path(),
            options=CatalogExportOptions(
                include_description_vi=True,
                include_description_en=True,
                preserve_image_order=False,
            ),
        )
        self.assertEqual(result.matched, 1)
        self.assertEqual(result.missing_codes, ())

    def test_preflight_save_updates_item_that_now_exists(self) -> None:
        item_id = self.create_item(
            code="77-777-77",
            description_vi="Old VI",
            description_en="Old EN",
        )
        issue = ExportPreflightItem(code="77-777-77", unknown_code=True)

        save_export_preflight_descriptions(
            issue,
            db=self.db,
            description_vi="Updated VI",
            description_en="Updated EN",
        )

        saved = self.db.get_item_by_code("77-777-77")
        self.assertEqual(saved.id, item_id)
        self.assertEqual(saved.description_vietnames_from_excel, "Updated VI")
        self.assertEqual(saved.description_excel, "Updated EN")

    def test_preflight_review_shows_pdf_english_fallback_without_copying_it(self) -> None:
        self.create_item(
            code="77-777-77",
            description_vi="",
            description_en="",
            pdf_description="Existing PDF English",
        )
        path = self.create_input_workbook([("77-777-77", 1)])
        rows = read_workbook_products(path)

        preflight = prepare_export_preflight(
            rows,
            db=self.db,
            include_description_vi=True,
            include_description_en=True,
        )
        issue = preflight.issues[0]

        self.assertTrue(issue.missing_vi)
        self.assertFalse(issue.missing_en)
        self.assertEqual(issue.description_en, "")
        self.assertEqual(issue.effective_description_en, "Existing PDF English")

        save_export_preflight_descriptions(
            issue,
            db=self.db,
            description_vi="New Vietnamese",
            description_en=issue.effective_description_en,
        )

        saved = self.db.get_item_by_code("77-777-77")
        self.assertEqual(saved.description_vietnames_from_excel, "New Vietnamese")
        self.assertEqual(saved.description_excel, "")
        self.assertEqual(saved.description, "Existing PDF English")
        self.assertFalse(issue.missing_vi)
        self.assertFalse(issue.missing_en)

    def test_reader_preflight_and_writer_are_callable_without_tkinter(self) -> None:
        self.create_item(
            code="12-345-67",
            description_vi="Vietnamese",
            description_en="English",
            image_specs=[("service.png", (40, 30), (20, 40, 60))],
        )
        path = self.create_input_workbook(
            [("12-345-67", 2), ("77-777-77", 5)],
            include_old_output=True,
        )

        rows = read_workbook_products(path)
        self.assertEqual(
            [(row.code, row.quantity) for row in rows],
            [("12-345-67", "2"), ("77-777-77", "5")],
        )
        preflight = prepare_export_preflight(
            rows,
            db=self.db,
            include_description_vi=True,
            include_description_en=True,
        )
        self.assertFalse(preflight.issues[0].has_issue)
        self.assertTrue(preflight.issues[1].unknown_code)

        result = export_catalog(
            path,
            preflight.rows,
            db=self.db,
            project_dir=self.project.path(),
            options=CatalogExportOptions(
                include_description_vi=True,
                include_description_en=True,
                preserve_image_order=True,
            ),
        )

        self.assertEqual(result.matched, 1)
        self.assertEqual(result.total, 2)
        self.assertEqual(result.images_found, 1)
        self.assertEqual(result.missing_codes, ("77-777-77",))
        workbook = load_workbook(path)
        try:
            self.assertEqual(workbook.sheetnames, ["Input", DEFAULT_SHEET_NAME])
            self.assertEqual(workbook[DEFAULT_SHEET_NAME]["B2"].value, "12-345-67")
        finally:
            workbook.close()


class PostProcessingSheetCharacterizationTests(ExportTestCase):
    def test_sheet_layout_highlighting_merges_and_replacement(self) -> None:
        workbook = Workbook()
        workbook.active.title = "Input"
        existing = workbook.create_sheet(DEFAULT_SHEET_NAME)
        existing["A1"] = "old"

        sheet = write_post_processing_sheet(
            workbook,
            [
                PostProcessingRow(
                    number=1,
                    code="12-345-67",
                    qty="2",
                    desc_primary="Mô tả VI",
                    desc_secondary="English",
                    force_secondary_row=True,
                ),
                PostProcessingRow(
                    number=2,
                    code="98-765-43",
                    qty="1",
                    desc_primary="",
                    desc_secondary="",
                    highlight_missing_desc=True,
                    force_secondary_row=True,
                    highlight_missing_secondary=True,
                ),
            ],
            index=1,
        )

        self.assertEqual(workbook.sheetnames, ["Input", DEFAULT_SHEET_NAME])
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet.print_title_rows, "$1:$1")
        self.assertEqual(sheet.page_setup.orientation, "portrait")
        self.assertEqual(
            [sheet.cell(1, column).value for column in range(1, 5)],
            ["No.", "Product Code", "Product Description", "Qty."],
        )
        self.assertEqual(
            [sheet.cell(2, column).value for column in range(1, 5)],
            [1, "12-345-67", "Mô tả VI", "2"],
        )
        self.assertEqual(sheet["C3"].value, "English")
        self.assertIn("A4:D4", {str(rng) for rng in sheet.merged_cells.ranges})
        self.assertEqual(sheet["C5"].fill.fgColor.rgb, "00FFFF00")
        self.assertEqual(sheet["C6"].fill.fgColor.rgb, "00FFFF00")
        self.assertIn("A7:D7", {str(rng) for rng in sheet.merged_cells.ranges})
        workbook.close()

    def test_image_loading_rotation_sorting_and_preserved_order(self) -> None:
        portrait = create_image_fixture(
            self.project.path("portrait.png"),
            size=(20, 60),
            color=(200, 20, 20),
        )
        landscape = create_image_fixture(
            self.project.path("landscape.png"),
            size=(80, 30),
            color=(20, 20, 200),
        )

        preserved = _load_images(
            [str(landscape), str(portrait)],
            preserve_image_order=True,
        )
        transformed = _load_images(
            [str(landscape), str(portrait)],
            preserve_image_order=False,
        )
        try:
            self.assertEqual(
                [(width, height) for _image, width, height in preserved],
                [(80, 30), (20, 60)],
            )
            # Non-preserved mode rotates portrait images, then sorts by area.
            self.assertEqual(
                [(width, height) for _image, width, height in transformed],
                [(60, 20), (80, 30)],
            )
        finally:
            for image, _width, _height in preserved + transformed:
                image.close()


class FullExportWorkflowCharacterizationTests(ExportTestCase):
    def test_headless_export_matches_codes_descriptions_images_and_quantities(self) -> None:
        self.create_item(
            code="12-345-67",
            description_vi="VI exact",
            description_en="EN exact",
            image_specs=[
                ("large.png", (80, 30), (200, 20, 20)),
                ("small.png", (30, 20), (20, 20, 200)),
            ],
        )
        self.create_item(
            code="98-765-43",
            description_vi="VI normalized",
            description_en="",
            pdf_description="PDF English fallback",
            image_specs=[("portrait.png", (20, 60), (20, 160, 20))],
        )
        path = self.create_input_workbook(
            [
                ("12-345-67", 2),
                ("98 – 765 – 43", "3"),
            ],
            include_old_output=True,
        )

        results, statuses, issues = self.run_export(
            path,
            include_vi=True,
            include_en=True,
            preserve_image_order=True,
        )

        self.assertEqual(issues, [])
        self.assertEqual(len(results), 1)
        self.assertEqual(
            {
                key: results[0][key]
                for key in (
                    "matched",
                    "total",
                    "images_found",
                    "missing_vi",
                    "missing_en",
                    "missing_images",
                    "missing_codes",
                )
            },
            {
                "matched": 2,
                "total": 2,
                "images_found": 2,
                "missing_vi": 0,
                "missing_en": 0,
                "missing_images": 0,
                "missing_codes": [],
            },
        )
        self.assertTrue(statuses)

        workbook = load_workbook(path)
        try:
            self.assertEqual(workbook.sheetnames, ["Input", DEFAULT_SHEET_NAME])
            output = workbook[DEFAULT_SHEET_NAME]
            self.assertEqual(
                [output.cell(2, column).value for column in range(1, 5)],
                [1, "12-345-67", "VI exact", "2"],
            )
            self.assertEqual(output["C3"].value, "EN exact")
            self.assertEqual(
                [output.cell(5, column).value for column in range(1, 5)],
                [2, "98 – 765 – 43", "VI normalized", "3"],
            )
            self.assertEqual(output["C6"].value, "PDF English fallback")
            self.assertEqual(len(output._images), 3)
            self.assertEqual(
                sorted(image.anchor._from.row + 1 for image in output._images),
                [4, 4, 7],
            )
        finally:
            workbook.close()

    def test_preflight_cancel_leaves_workbook_unchanged(self) -> None:
        self.create_item(
            code="12-345-67",
            description_vi="",
            description_en="",
        )
        path = self.create_input_workbook([("12-345-67", 1)])
        before_package = snapshot_xlsx_package(path)

        results, statuses, issues = self.run_export(
            path,
            include_vi=True,
            include_en=True,
            preserve_image_order=False,
            review_action="cancel",
        )

        self.assertEqual(results, [])
        self.assertEqual(len(issues), 1)
        self.assertTrue(issues[0][0].missing_vi)
        self.assertTrue(issues[0][0].missing_en)
        self.assertTrue(issues[0][0].missing_images)
        self.assertEqual(snapshot_xlsx_package(path), before_package)
        self.assertTrue(any("y" in status.lower() or status for status in statuses))

    def test_continue_exports_missing_and_unknown_rows_with_highlighting(self) -> None:
        self.create_item(
            code="12-345-67",
            description_vi="",
            description_en="Known English",
        )
        path = self.create_input_workbook(
            [
                ("12-345-67", 1),
                ("77-777-77", 4),
            ]
        )

        results, _statuses, issues = self.run_export(
            path,
            include_vi=True,
            include_en=True,
            preserve_image_order=False,
            review_action="continue",
        )

        self.assertEqual(len(issues), 1)
        self.assertEqual(len(issues[0]), 2)
        by_code = {issue.code: issue for issue in issues[0]}
        self.assertTrue(by_code["12-345-67"].missing_vi)
        self.assertTrue(by_code["12-345-67"].missing_images)
        self.assertTrue(by_code["77-777-77"].unknown_code)
        self.assertEqual(results[0]["matched"], 1)
        self.assertEqual(results[0]["total"], 2)
        self.assertEqual(results[0]["missing_vi"], 1)
        self.assertEqual(results[0]["missing_images"], 1)
        self.assertEqual(results[0]["missing_codes"], ["77-777-77"])

        workbook = load_workbook(path)
        try:
            output = workbook[DEFAULT_SHEET_NAME]
            self.assertEqual(output["C2"].fill.fgColor.rgb, "00FFFF00")
            self.assertEqual(output["C3"].value, "Known English")
            self.assertEqual(output["C5"].fill.fgColor.rgb, "00FFFF00")
            self.assertEqual(output["B5"].value, "77-777-77")
        finally:
            workbook.close()

    def test_vi_only_and_en_only_options_keep_current_row_shapes(self) -> None:
        self.create_item(
            code="12-345-67",
            description_vi="Vietnamese",
            description_en="English",
            image_specs=[("item.png", (40, 30), (60, 80, 100))],
        )

        vi_path = self.create_input_workbook([("12-345-67", 1)])
        self.run_export(
            vi_path,
            include_vi=True,
            include_en=False,
            preserve_image_order=False,
        )
        vi_snapshot = snapshot_workbook(vi_path)
        vi_sheet = vi_snapshot["sheets"][1]
        self.assertIn(["C2", "Vietnamese"], vi_sheet["cells"])
        self.assertIn("A3:D3", vi_sheet["merged_ranges"])

        en_path = self.project.path("export-en.xlsx")
        vi_path.replace(en_path)
        # Recreate input because the previous call has already produced output.
        workbook = load_workbook(en_path)
        try:
            input_sheet = workbook["Input"]
            while len(workbook.worksheets) > 1:
                workbook.remove(workbook.worksheets[-1])
            workbook.save(en_path)
        finally:
            workbook.close()
        self.run_export(
            en_path,
            include_vi=False,
            include_en=True,
            preserve_image_order=False,
        )
        en_snapshot = snapshot_workbook(en_path)
        en_sheet = en_snapshot["sheets"][1]
        self.assertIn(["C2", "English"], en_sheet["cells"])
        self.assertIn("A3:D3", en_sheet["merged_ranges"])


class BrandingPackageCharacterizationTests(ExportTestCase):
    def create_branding(self) -> Path:
        return create_image_fixture(
            self.data_dir / "assets" / "bg.jpg",
            size=(200, 40),
            color=(30, 80, 140),
        )

    def create_formatted_workbook(self) -> Path:
        path = self.project.path("branded.xlsx")
        workbook = Workbook()
        workbook.active.title = "Input"
        write_post_processing_sheet(
            workbook,
            [
                PostProcessingRow(
                    number=1,
                    code="12-345-67",
                    desc_primary="Description",
                )
            ],
        )
        workbook.save(path)
        workbook.close()
        return path

    def test_missing_branding_is_a_noop(self) -> None:
        path = self.create_formatted_workbook()
        before = snapshot_xlsx_package(path)

        self.assertFalse(
            apply_post_processing_branding(
                path,
                project_dir=self.project.path(),
            )
        )
        self.assertEqual(snapshot_xlsx_package(path), before)

    def test_branding_injects_vml_relationships_media_and_header_footer(self) -> None:
        self.create_branding()
        path = self.create_formatted_workbook()

        self.assertTrue(
            apply_post_processing_branding(
                path,
                project_dir=self.project.path(),
            )
        )

        with zipfile.ZipFile(path, "r") as archive:
            names = set(archive.namelist())
            vml_paths = sorted(
                name
                for name in names
                if name.startswith("xl/drawings/vmlDrawing") and name.endswith(".vml")
            )
            vml_rel_paths = sorted(
                name
                for name in names
                if name.startswith("xl/drawings/_rels/vmlDrawing")
            )
            jpeg_paths = sorted(
                name
                for name in names
                if name.startswith("xl/media/") and name.endswith(".jpeg")
            )
            sheet_xml = archive.read("xl/worksheets/sheet2.xml").decode("utf-8")
            sheet_rels = archive.read(
                "xl/worksheets/_rels/sheet2.xml.rels"
            ).decode("utf-8")
            content_types = archive.read("[Content_Types].xml").decode("utf-8")

        self.assertEqual(len(vml_paths), 1)
        self.assertEqual(len(vml_rel_paths), 1)
        self.assertEqual(len(jpeg_paths), 1)
        self.assertIn("legacyDrawingHF", sheet_xml)
        self.assertIn("headerFooter", sheet_xml)
        self.assertIn("vmlDrawing", sheet_rels)
        self.assertIn('Extension="vml"', content_types)
        self.assertIn('Extension="jpeg"', content_types)

        workbook = load_workbook(path)
        try:
            self.assertEqual(workbook[DEFAULT_SHEET_NAME].oddHeader.center.text, "&G")
        finally:
            workbook.close()

    def test_reapplying_branding_replaces_old_header_assets(self) -> None:
        self.create_branding()
        path = self.create_formatted_workbook()
        apply_post_processing_branding(path, project_dir=self.project.path())
        apply_post_processing_branding(path, project_dir=self.project.path())

        with zipfile.ZipFile(path, "r") as archive:
            names = archive.namelist()
        self.assertEqual(
            len(
                [
                    name
                    for name in names
                    if name.startswith("xl/drawings/vmlDrawing")
                    and name.endswith(".vml")
                ]
            ),
            1,
        )
        self.assertEqual(
            len(
                [
                    name
                    for name in names
                    if name.startswith("xl/media/") and name.endswith(".jpeg")
                ]
            ),
            1,
        )


if __name__ == "__main__":
    unittest.main()
